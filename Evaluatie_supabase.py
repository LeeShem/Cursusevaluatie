import streamlit as st
import pandas as pd
import matplotlib.pyplot as plt
import json, os, io, re, base64
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from datetime import datetime
from supabase import create_client

# ═══════════════════════════════════════════════════════════════
#  CONFIG  –  pas hier aan per cursus
# ═══════════════════════════════════════════════════════════════
DOCENT_WACHTWOORD = "alo"           # ← wachtwoord voor docentendashboard

# Studiehandleiding PDF
# Zet het PDF-bestand in dezelfde map als dit script en pas de naam hieronder aan.
STUDIEHANDLEIDING_PAD  = "SHL.pdf"      # ← bestandsnaam aanpassen
STUDIEHANDLEIDING_NAAM = "Studiehandleiding cursus verantwoord leren lesgeven"   # ← weergavenaam aanpassen

# ═══════════════════════════════════════════════════════════════
#  SUPABASE CONFIG
#  De verbindingsgegevens worden ingelezen via Streamlit secrets.
#  Je hoeft hier niets aan te passen.
# ═══════════════════════════════════════════════════════════════
TABEL_ST = "studenten_resultaten"   # naam tabel studenten in Supabase
TABEL_WV = "werkveld_resultaten"    # naam tabel werkveld in Supabase
TABEL_DC = "docent_resultaten"      # naam tabel docent eigen evaluatie in Supabase

@st.cache_resource
def verbind_supabase():
    url = st.secrets["SUPABASE_URL"]
    key = st.secrets["SUPABASE_KEY"]
    return create_client(url, key)

def laad(tabel: str) -> list:
    try:
        sb = verbind_supabase()
        response = sb.table(tabel).select("data").execute()
        resultaten = []
        for rij in response.data:
            try:
                resultaten.append(json.loads(rij["data"]))
            except Exception:
                pass
        return resultaten
    except Exception as e:
        st.error(f"Fout bij laden van data: {e}")
        return []

def schrijf_rij(tabel: str, data: dict):
    try:
        sb = verbind_supabase()
        sb.table(tabel).insert({"data": json.dumps(data, ensure_ascii=False)}).execute()
    except Exception as e:
        st.error(f"Fout bij opslaan: {e}")

def verwijder_alle(tabel: str):
    try:
        sb = verbind_supabase()
        sb.table(tabel).delete().neq("id", 0).execute()
    except Exception as e:
        st.error(f"Fout bij verwijderen: {e}")

# ═══════════════════════════════════════════════════════════════
#  VRAGEN STUDENTEN
# ═══════════════════════════════════════════════════════════════
VRAGEN_ST = {
    "Instructie": [
        ("In welke mate draagt de inhoud van de les bij aan het behalen van het lesdoel?",               1),
        ("In hoeverre worden de lesdoelen aan het begin van elke les duidelijk gecommuniceerd?",          2),
        ("De lesstof wordt verbonden aan realistische situaties vanuit de stagepraktijk.",                 2),
        ("De lesstof wordt toegelicht waarom het relevant is in de praktijk?",                            3),
        ("De lesstof wordt op een behapbare wijze aangeboden.",                                           3),
        ("De verplichtingen en verwachtingen van de cursus worden duidelijk gecommuniceerd.",             3),
        ("De visuele hulpmiddelen helpen met het begrip van de leerstof.",                                4),
    ],
    "Verwerking": [
        ("De verwerkingsopdracht sluit aan op de behandelde stof.",                                       [1,2,3,4]),
        ("De verwerkingsopdracht stimuleert je om de lesstof actief te verwerken.",                      [1,2,3,4]),
        ("De verwerkingsopdracht wordt gekoppeld aan actuele praktijkproblemen.",                         [1,2,3,4]),
    ],
    "Studieoverzicht": [
        ("De hoeveelheid stof is in verhouding tot het aantal beschikbare lessen.",                       1),
        ("Het studiemateriaal is op een overzichtelijke wijze terug te vinden.",                          4),
        ("De docent geeft duidelijk aan waar de stof terug te vinden is.",                                4),
        ("De docent begeleidt je bij het terugvinden van de leerstof wanneer dit niet lukt.",             4),
        ("Er is een duidelijk overzicht waaruit blijkt welke stof op welk moment wordt behandeld.",       4),
    ],
    "Docent": [
        ("De docent is beschikbaar en bereikbaar voor het beantwoorden van vragen.",   None),
        ("De docent heeft aandacht voor jou als individu binnen de groep.",             None),
        ("De docent maakt de stof toegankelijk en begrijpelijk.",                       None),
        ("De docent denkt actief met je mee bij het oplossen van eventuele problemen.", None),
    ],
    "Toetsing": [
        ("De formatieve toetsing geeft je een duidelijk beeld wat je kunt verwachten bij de summatieve toets.", 1),
        ("In hoeverre zijn de beoordelingscriteria helder en transparant?",                               2),
        ("In hoeverre dekt de toets de behandelde stof?",                                                 2),
        ("In hoeverre bevraagt de toetsvorm je op het toepassen van kennis in realistische praktijksituaties?", 3),
        ("In welke mate maakt de formatieve toetsing inzichtelijk aan welke onderdelen je nog verder moet werken?", 4),
    ],
    "Overig": [
        ("Andere punten die u wilt delen (tips, tricks etc.) wat ten goede komt aan de cursus.", "open"),
        ("Heb je de cursus gehaald?", "cursus_gehaald"),
    ],
}

RUBRIC_ONDERDELEN = ["Instructie", "Verwerking", "Studieoverzicht", "Toetsing"]

RUBRIC_INHOUD = {
    ("Instructie", 1): "De behandelde theorie tijdens de instructie is direct te herleiden naar de formele leerdoelen en/of de eindtoets.\n\nDe docent stelt tijdens de instructie minimaal één gerichte denkvraag (geen reproductie) aan de groep.",
    ("Instructie", 2): "De les start expliciet met het benoemen van het leerdoel en het activeren van voorkennis.\n\nDe theorie-uitleg begint met een herkenbaar praktijkprobleem uit de gymzaal.\n\nDe lesstof wordt verbonden aan realistische situaties vanuit de stagepraktijk.",
    ("Instructie", 3): "De docent benoemt expliciet de relevantie van de theorie voor de latere rol als LO-docent.\n\nDe lesstof wordt op een behapbare wijze aangeboden.\n\nDe verplichtingen en verwachtingen van de cursus worden duidelijk gecommuniceerd.",
    ("Instructie", 4): "De inzet van media is doelbewust: het ondersteunt de theorie of dient als bewuste brain break, maar leidt niet af tijdens complexe uitleg.\n\nDe docent benoemt het onderscheid tussen de hoofdzaken en de bijzaken.",
    ("Verwerking", 1): "De opdracht vereist expliciet het gebruik van academische hbo-vaardigheden.\n\nDe verwerkingsopdracht sluit aan op de behandelde stof.",
    ("Verwerking", 2): "De werkvorm dwingt studenten om zelf verbanden te zoeken en betekenis te geven aan de stof (Deep Learning).\n\nDe verwerkingsopdracht stimuleert actieve verwerking van de lesstof.",
    ("Verwerking", 3): "De opdracht dwingt de student om de theorie daadwerkelijk uit te voeren of te simuleren.\n\nDe verwerkingsopdracht wordt gekoppeld aan actuele praktijkproblemen.",
    ("Verwerking", 4): "De opdracht biedt de student bewust ruimte voor eigen inbreng, keuzes of een eigen leerweg.\n\nEr is een duidelijk moment ingericht voor de ontwikkeling van zelfregulatie.",
    ("Studieoverzicht", 1): "De totale hoeveelheid te bestuderen stof is realistisch gecalculeerd in verhouding tot de formele studiebelastingsuren (SBU's).",
    ("Studieoverzicht", 2): "Geen interessante informatie van de studenten verkrijgbaar",
    ("Studieoverzicht", 3): "Geen interessante informatie van de studenten verkrijgbaar",
    ("Studieoverzicht", 4): "Het materiaal is in een bruikbaar digitaal formaat beschikbaar gesteld.\n\nEr is een helder overzicht (planning) beschikbaar.\n\nDe docent geeft duidelijk aan waar de stof terug te vinden is.",
    ("Toetsing", 1): "De inhoud, vraagstelling en moeilijkheidsgraad van de formatieve oefenoets(en) komen aantoonbaar overeen met het summatieve eindtentamen.",
    ("Toetsing", 2): "De toetsvragen maken structureel gebruik van uitgewerkte praktijkcasussen.\n\nHet scoringsmodel beloont expliciet diepgaand begrip.\n\nDe toetsmatrijs toont een bewuste balans.",
    ("Toetsing", 3): "De gekozen toetsvorm is valide voor het leerdoel: handelscompetenties worden getoetst via een praktijk assessment.\n\nDe casus in de toets is een authentieke weergave van een probleem uit de actuele ALO-beroepspraktijk.",
    ("Toetsing", 4): "Bij formatieve evaluaties krijgt de student naast een oordeel ook concrete feed forward (handelingsperspectief).",
}

# ═══════════════════════════════════════════════════════════════
#  VRAGEN WERKVELD
# ═══════════════════════════════════════════════════════════════
VRAGEN_WV = {
    "A - Inhoud & Relevantie": [
        "Sluiten de inhoud vanuit de cursus aan bij de situaties die u tegenkomt in de praktijk?",
        "De leerdoelen waren duidelijk en relevant voor de beroepspraktijk.",
        "De diepgang van de cursus was passend op de beginsituatie van de stage.",
        "In welke mate vindt u de cursus goed aansluiten op de stage?",
    ],
    "B - Toepasbaarheid in de praktijk": [
        "In welke mate ziet u dat de student de leerdoelen uit de cursus toepast in de praktijk?",
        "In welke mate ziet u dat de student de geleerde inhoud toepast binnen de praktijk?",
    ],
}


# ═══════════════════════════════════════════════════════════════
#  VRAGEN DOCENT EIGEN EVALUATIE
#  Elke stelling: (tekst, lens_nr, theorie_label)
# ═══════════════════════════════════════════════════════════════
VRAGEN_DC = {
    "1 - Instructie": [
        ("De behandelde theorie tijdens de instructie is direct te herleiden naar de formele leerdoelen en/of de eindtoets (geen irrelevante informatie).", 1, "Biggs"),
        ("De docent stelt tijdens de instructie minimaal één gerichte denkvraag (geen reproductie) aan de groep.", 1, "Bloom"),
        ("De les start expliciet met het benoemen van het leerdoel en het activeren van voorkennis.", 2, "Gagné"),
        ("De theorie-uitleg begint met een herkenbaar praktijkprobleem uit de gymzaal (i.p.v. droge definities).", 2, "Context-Concept"),
        ("Complexe theorie wordt op de slides in visueel toegankelijke stappen (chunks) aangeboden.", 2, "Scaffolding"),
        ("De docent benoemt expliciet de relevantie van de theorie voor de latere rol als LO-docent.", 3, "Shulman PCK"),
        ("De docent benoemt tijdens de uitleg expliciet het onderscheid tussen de hoofdzaken (de kern) en de bijzaken.", 4, "Zimmerman"),
        ("De inzet van media is doelbewust: het ondersteunt de theorie óf dient als bewuste brain break, maar leidt niet af tijdens complexe uitleg.", 4, "TPACK"),
    ],
    "2 - Verwerking": [
        ("De opdracht vereist expliciet het gebruik van academische hbo-vaardigheden (zoals bronnen wegen, professioneel argumenteren of kritisch analyseren).", 1, "Dublin"),
        ("Het zwaartepunt van de opdracht ligt aantoonbaar op Toepassen of Evalueren (de student moet iets doen met de kennis, niet slechts reproduceren).", 1, "Bloom"),
        ("De werkvorm dwingt studenten om zelf verbanden te zoeken en betekenis te geven aan de stof (Deep Learning), in plaats van passief invulwerk te doen.", 2, "Entwistle"),
        ("De studenten werken actief en interactief (bijv. via discussie of ontwerp) aan een betekenisvol praktijkprobleem.", 2, "Context-Concept"),
        ("Tijdens het werkcollege neemt de expliciete sturing van de docent merkbaar af, waardoor de autonomie van de student toeneemt.", 2, "Scaffolding"),
        ("De opdracht dwingt de student om de theorie daadwerkelijk uit te voeren of te simuleren (Shows How), in plaats van er alleen over te praten (Knows).", 3, "Miller"),
        ("De theorie wordt door de student actief gekoppeld aan een compleet nieuwe situatie of casus, waarmee transfer wordt geoefend.", 3, "Kolb"),
        ("De opdracht of de docent biedt de student bewust ruimte voor eigen inbreng, keuzes of een eigen leerweg (autonomie).", 4, "Zimmerman"),
        ("Er is een duidelijk moment in de les ingericht voor de ontwikkeling van zelfregulatie (zoals het geven van peer-feedback of reflectie op het eigen proces).", 4, "Zimmerman"),
    ],
    "3 - Zelfstudie & Literatuur": [
        ("De teksten sluiten qua taal- en denkniveau aan bij de HBO ALO-student (geen versimpelde taal, maar ook geen onnodig complexe universitaire papers).", 1, "Dublin"),
        ("De totale hoeveelheid te bestuderen stof is realistisch gecalculeerd in verhouding tot de formele studiebelastingsuren (SBU's).", 1, "Biggs"),
        ("De literatuur wordt niet zomaar opgegeven, maar is voorzien van een heldere leesinstructie, focuspunten of specifieke leesvragen.", 2, "Scaffolding"),
        ("In de opdracht tot zelfstudie zit een trigger die de student dwingt om eerst zijn/haar voorkennis op te halen voordat het lezen begint.", 2, "Gagné"),
        ("Er staat expliciet beschreven waarom deze specifieke literatuur relevant is, om zo de diepe motivatie te stimuleren in plaats van vinkjesgedrag.", 2, "Entwistle"),
        ("De literatuur is geschreven voor (of direct en logisch te vertalen naar) de ALO-context, en is geen algemene psychologie of theorie zonder link naar de gymzaal.", 3, "Shulman PCK"),
        ("De leesopdracht vraagt de student expliciet om tijdens het lezen de theorie (concepten) te koppelen aan een eigen (stage)ervaring.", 3, "Kolb"),
        ("Het materiaal is in een bruikbaar digitaal formaat beschikbaar gesteld, zodat de student de tekst kan doorzoeken en digitaal kan markeren.", 4, "TPACK"),
        ("Er is een helder overzicht (planning) beschikbaar waaruit de student exact kan opmaken wanneer welke stof bestudeerd én toegepast moet zijn.", 4, "Zimmerman"),
    ],
    "4 - Digitale Tools & Oefenmateriaal": [
        ("De oefenvragen in de tool komen qua format en denkniveau aantoonbaar overeen met de vragen op het uiteindelijke tentamen (alignment).", 1, "Biggs"),
        ("De tool toetst niet uitsluitend op feitenkennis (wat), maar vraagt actief naar inzicht en het leggen van verbanden (waarom/hoe).", 1, "Bloom"),
        ("De digitale tool biedt de mogelijkheid tot differentiatie (bijv. snellere studenten krijgen automatisch moeilijkere of verdiepende vragen).", 2, "Entwistle"),
        ("De tool is logisch en zichtbaar ingebed in de leerroute van de student (als leermiddel), en staat niet als een losstaande, geïsoleerde gimmick in de cursus.", 2, "Gagné"),
        ("De tool wordt gebruikt om een realistische praktijkhandeling te simuleren (Shows how), waarmee de student de brug naar de praktijk kan oefenen.", 3, "Miller"),
        ("De tool schetst realistische praktijksituaties (ervaring) waarmee de student veilig en actief kan experimenteren zonder direct in de gymzaal te hoeven staan.", 3, "Kolb"),
        ("Bij een foutief antwoord of actie geeft de tool direct specifieke, inhoudelijke feedback (feed forward), in plaats van uitsluitend de melding goed/fout.", 4, "Zimmerman"),
        ("De ingezette digitale tool biedt een duidelijke didactische meerwaarde die zonder deze specifieke technologie niet (of zeer moeizaam) gerealiseerd had kunnen worden.", 4, "TPACK"),
    ],
    "5 - Beroepspraktijk": [
        ("De weging van de stage-opdracht is expliciet vastgelegd en telt (formatief of summatief) zichtbaar mee voor de afronding van de cursus.", 1, "Biggs"),
        ("De opdracht eist dat de student zelf een les(deel) ontwerpt of analytisch onderbouwt (waarom doe ik dit?), in plaats van uitsluitend een bestaand plan uit te voeren.", 1, "Bloom"),
        ("De stage-instructie bevat concrete focuspunten, kaders of een kijkwijzer, zodat het voor de student helder is op welke hoofdzaken gelet moet worden.", 2, "Scaffolding"),
        ("De opbouw van de opdracht dwingt de student om eerst in de gymzaal te handelen/observeren (context) voordat dit theoretisch verklaard wordt (concept).", 2, "Context-Concept"),
        ("De opdracht borgt in de stappen dat de volledige leercyclus wordt doorlopen: ervaring opdoen, reflecteren, theorie koppelen, en een nieuw plan maken.", 3, "Kolb"),
        ("Er is een verplicht, aantoonbaar moment (bijv. een handtekening of verslagje) waarop de student de theorie bespreekt met de werkplekbegeleider.", 3, "Kolb"),
        ("De instructie verplicht de student om voorafgaand aan de praktijkuitvoering expliciet eigen, persoonlijke leerdoelen te formuleren.", 4, "Zimmerman"),
        ("Het stage-portfolio of de opdracht bevat een verplicht onderdeel voor zowel tussentijdse als afrondende reflectie op het eigen leerproces.", 4, "Zimmerman"),
    ],
    "6 - Toetsing & Evaluatie": [
        ("De inhoud, vraagstelling en moeilijkheidsgraad van de formatieve oefentoets(en) komen aantoonbaar overeen met het summatieve eindtentamen.", 1, "Biggs"),
        ("De toetsmatrijs toont een bewuste, afgewogen balans tussen reproductievragen (kennis stampen) en hogere-orde vragen (toepassen, analyseren).", 1, "Bloom"),
        ("De toetsvragen (ook schriftelijk) maken structureel gebruik van uitgewerkte praktijkcasussen, in plaats van enkel te vragen naar droge theoretische definities.", 2, "Context-Concept"),
        ("Het scoringsmodel of de rubric beloont expliciet het aantonen van diepgaand begrip en het leggen van eigen verbanden (Deep Learning), en niet slechts exact reproductiewerk.", 2, "Entwistle"),
        ("De gekozen toetsvorm is valide voor het leerdoel: handelingscompetenties (zoals lesgeven) worden getoetst via een praktijkassessment (Shows how), niet via een kennistoets.", 3, "Miller"),
        ("Het probleem of de casus die in de toets centraal staat, is een authentieke, waarheidsgetrouwe weergave van een probleem uit de actuele ALO-beroepspraktijk.", 3, "Miller"),
        ("Bij formatieve evaluaties krijgt de student naast een oordeel ook concrete feed forward (handelingsperspectief), zodat duidelijk is wat er moet gebeuren om beter te worden.", 4, "Zimmerman"),
    ],
}

# Rubric inhoud per (onderdeel_key, lens) voor de docent evaluatie
RUBRIC_DC = {
    ("1 - Instructie", 1): "Biggs: De behandelde theorie is direct te herleiden naar de leerdoelen.\n\nBloom: De docent stelt minimaal één gerichte denkvraag aan de groep.",
    ("1 - Instructie", 2): "Gagné: De les start met benoemen leerdoel en activeren voorkennis.\n\nContext-Concept: Uitleg begint met herkenbaar praktijkprobleem.\n\nScaffolding: Complexe theorie in toegankelijke stappen aangeboden.",
    ("1 - Instructie", 3): "Shulman PCK: De docent benoemt expliciet de relevantie voor de rol als LO-docent.",
    ("1 - Instructie", 4): "Zimmerman: Onderscheid hoofdzaken en bijzaken wordt benoemd.\n\nTPACK: Media-inzet is doelbewust en leidt niet af.",
    ("2 - Verwerking", 1): "Dublin: Opdracht vereist gebruik van academische hbo-vaardigheden.\n\nBloom: Zwaartepunt ligt op Toepassen of Evalueren.",
    ("2 - Verwerking", 2): "Entwistle: Werkvorm dwingt studenten verbanden te zoeken (Deep Learning).\n\nContext-Concept: Actief en interactief werken aan praktijkprobleem.\n\nScaffolding: Sturing docent neemt af, autonomie student neemt toe.",
    ("2 - Verwerking", 3): "Miller: Opdracht dwingt theorie uit te voeren of simuleren (Shows How).\n\nKolb: Theorie actief koppelen aan nieuwe situatie (transfer).",
    ("2 - Verwerking", 4): "Zimmerman: Ruimte voor eigen inbreng en autonomie.\n\nZimmerman: Moment ingericht voor zelfregulatie (peer-feedback of reflectie).",
    ("3 - Zelfstudie & Literatuur", 1): "Dublin: Teksten sluiten aan bij HBO ALO-denkniveau.\n\nBiggs: Hoeveelheid stof realistisch in verhouding tot SBU's.",
    ("3 - Zelfstudie & Literatuur", 2): "Scaffolding: Literatuur voorzien van leesinstructie of focuspunten.\n\nGagné: Trigger om voorkennis op te halen voor het lezen.\n\nEntwistle: Beschrijving waarom literatuur relevant is.",
    ("3 - Zelfstudie & Literatuur", 3): "Shulman PCK: Literatuur direct vertaalbaar naar ALO-context.\n\nKolb: Student koppelt theorie aan eigen (stage)ervaring.",
    ("3 - Zelfstudie & Literatuur", 4): "TPACK: Materiaal in bruikbaar digitaal formaat beschikbaar.\n\nZimmerman: Helder overzicht wanneer welke stof bestudeerd moet zijn.",
    ("4 - Digitale Tools & Oefenmateriaal", 1): "Biggs: Oefenvragen komen overeen met tentamen (alignment).\n\nBloom: Tool vraagt actief naar inzicht, niet alleen feitenkennis.",
    ("4 - Digitale Tools & Oefenmateriaal", 2): "Entwistle: Tool biedt mogelijkheid tot differentiatie.\n\nGagné: Tool is logisch ingebed in de leerroute.",
    ("4 - Digitale Tools & Oefenmateriaal", 3): "Miller: Tool simuleert realistische praktijkhandeling (Shows how).\n\nKolb: Tool schetst realistische praktijksituaties voor experimenteren.",
    ("4 - Digitale Tools & Oefenmateriaal", 4): "Zimmerman: Tool geeft specifieke feed forward bij foutief antwoord.\n\nTPACK: Tool biedt duidelijke didactische meerwaarde.",
    ("5 - Beroepspraktijk", 1): "Biggs: Weging stage-opdracht expliciet vastgelegd.\n\nBloom: Student ontwerpt zelf een les of onderbouwt analytisch.",
    ("5 - Beroepspraktijk", 2): "Scaffolding: Instructie bevat concrete focuspunten en kijkwijzer.\n\nContext-Concept: Eerst handelen/observeren, dan theoretisch verklaren.",
    ("5 - Beroepspraktijk", 3): "Kolb: Volledige leercyclus wordt doorlopen.\n\nKolb: Verplicht moment waarop student theorie bespreekt met begeleider.",
    ("5 - Beroepspraktijk", 4): "Zimmerman: Student formuleert vooraf persoonlijke leerdoelen.\n\nZimmerman: Portfolio bevat tussentijdse en afrondende reflectie.",
    ("6 - Toetsing & Evaluatie", 1): "Biggs: Formatieve oefentoets komt overeen met summatief eindtentamen.\n\nBloom: Toetsmatrijs toont balans reproductie en hogere-orde vragen.",
    ("6 - Toetsing & Evaluatie", 2): "Context-Concept: Toetsvragen maken gebruik van praktijkcasussen.\n\nEntwistle: Scoringsmodel beloont diepgaand begrip (Deep Learning).",
    ("6 - Toetsing & Evaluatie", 3): "Miller: Toetsvorm valide voor leerdoel (Shows how bij handelingscompetenties).\n\nMiller: Casus is authentieke weergave van ALO-beroepspraktijk.",
    ("6 - Toetsing & Evaluatie", 4): "Zimmerman: Student krijgt naast oordeel ook concrete feed forward.",
}

SCORE_LABELS_DC = {1: "Niet aanwezig", 2: "Beperkt aanwezig", 3: "Naar wens aanwezig"}
SCORE_NVT = "NVT"
SCORE_KLEUREN_DC = {1: "#e74c3c", 2: "#f1c40f", 3: "#2ecc71"}

def dc_kleur(gem):
    """Kleur op basis van gemiddelde score 1-3."""
    if gem is None: return "#e8ecf4"
    if gem < 1.67: return "#f8d7da"   # rood
    elif gem < 2.34: return "#fff9c4"  # geel
    else: return "#d4edda"             # groen

def dc_kleur_hex(gem):
    if gem is None: return "E8ECF4"
    if gem < 1.67: return "F8D7DA"
    elif gem < 2.34: return "FFF9C4"
    else: return "D4EDDA"

def dc_kleur_tekst(gem):
    if gem is None: return "#888"
    if gem < 1.67: return "#721c24"
    elif gem < 2.34: return "#533f03"
    else: return "#155724"

def dc_label(gem):
    """Geeft het label terug op basis van gemiddelde score (NVT uitgesloten)."""
    if gem is None: return "Geen data"
    if gem < 1.67: return "Niet aanwezig"
    elif gem < 2.34: return "Beperkt aanwezig"
    else: return "Naar wens aanwezig"

# ═══════════════════════════════════════════════════════════════
#  GEDEELDE CONSTANTEN
# ═══════════════════════════════════════════════════════════════
NIVEAU_LABELS = {
    1: "Niveau 1 - Heel slecht", 2: "Niveau 2 - Slecht",
    3: "Niveau 3 - Voldoende",   4: "Niveau 4 - Goed",
    5: "Niveau 5 - Excellent",
}
NIVEAU_KLEUREN = {1:"#e74c3c",2:"#e67e22",3:"#f1c40f",4:"#2ecc71",5:"#1abc9c"}
NIVEAU_BESCHRIJVING = {
    1:"De cursus werd als heel slecht ervaren. Er zijn grote verbeterpunten.",
    2:"De cursus werd als slecht ervaren. Meerdere aspecten verdienen aandacht.",
    3:"De cursus werd als voldoende ervaren. Er is ruimte voor verbetering.",
    4:"De cursus werd als goed ervaren. Kleine verbeteringen zijn nog mogelijk.",
    5:"De cursus werd als excellent ervaren. Uitstekend resultaat!",
}

SCHAAL_INFO = """<div class="schaal-info">
<h4>Info Beoordelingsschaal</h4>
<p><strong>De beoordelingscategorieen zijn als volgt ingedeeld:</strong></p>
<ul>
    <li><strong>Zeer slecht:</strong> 1</li>
    <li><strong>Slecht:</strong> 2</li>
    <li><strong>Neutraal:</strong> 3</li>
    <li><strong>Goed:</strong> 4</li>
    <li><strong>Zeer goed:</strong> 5</li>
</ul>
<p>Deze indeling helpt om duidelijk onderscheid te maken tussen verschillende niveaus en
te kijken waar de verbeterpunten liggen.</p>
</div>"""

# ═══════════════════════════════════════════════════════════════
#  KLEUR HELPERS
# ═══════════════════════════════════════════════════════════════
def rubric_kleur(g):
    if g is None: return "#e8ecf4"
    if g < 2.0:   return "#f8d7da"
    elif g < 3.0: return "#fde8c8"
    elif g < 4.0: return "#fff9c4"
    elif g < 4.5: return "#d4edda"
    else:         return "#a8e6cf"

def rubric_kleur_tekst(g):
    if g is None: return "#888"
    if g < 2.0:   return "#721c24"
    elif g < 3.0: return "#856404"
    elif g < 4.0: return "#533f03"
    elif g < 4.5: return "#155724"
    else:         return "#0a4a2a"

def rubric_hex(g):
    if g is None: return "E8ECF4"
    if g < 2.0:   return "F8D7DA"
    elif g < 3.0: return "FDE8C8"
    elif g < 4.0: return "FFF9C4"
    elif g < 4.5: return "D4EDDA"
    else:         return "A8E6CF"

def niveau_kleur_css(n): return NIVEAU_KLEUREN.get(n, "#95a5a6")

# ═══════════════════════════════════════════════════════════════
#  DATA HELPERS
# ═══════════════════════════════════════════════════════════════
def bereken_niveau(g):
    if g < 1.75: return 1
    elif g < 2.5: return 2
    elif g < 3.5: return 3
    elif g < 4.5: return 4
    else: return 5

def is_geldig_email(e):
    return bool(re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", e.strip()))



def sla_student_op(spv, sg, sn, tn, open_antwoord="", cursus_gehaald=""):
    schrijf_rij(TABEL_ST, {
        "tijdstip": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "scores_per_vraag": spv,
        "sectie_gemiddeldes": sg,
        "sectie_niveaus": sn,
        "totaal_niveau": tn,
        "open_antwoord": open_antwoord.strip(),
        "cursus_gehaald": cursus_gehaald,
    })

def sla_werkveld_op(email, scores, spv, niveaus, tn, fg):
    schrijf_rij(TABEL_WV, {
        "email": email,
        "tijdstip": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "scores": scores,
        "scores_per_vraag": spv,
        "niveaus": niveaus,
        "totaal_niveau": tn,
        "focusgroep": fg,
    })

def sla_docent_op(scores_per_stelling, sectie_gemiddeldes, sectie_niveaus, totaal_gem, argumentaties=None):
    schrijf_rij(TABEL_DC, {
        "tijdstip": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "scores_per_stelling": scores_per_stelling,
        "sectie_gemiddeldes": sectie_gemiddeldes,
        "sectie_niveaus": sectie_niveaus,
        "totaal_gemiddelde": totaal_gem,
        "argumentaties": argumentaties or {},
    })

def bereken_lens_gemiddeldes(resultaten):
    ls = {(o, l): [] for o in RUBRIC_ONDERDELEN for l in [1,2,3,4]}
    for r in resultaten:
        for o, vl in VRAGEN_ST.items():
            if o not in RUBRIC_ONDERDELEN: continue
            sc = r.get("scores_per_vraag", {}).get(o, [])
            for i, (_, lens_val) in enumerate(vl):
                if lens_val is None or lens_val == "open" or i >= len(sc): continue
                lensen = lens_val if isinstance(lens_val, list) else [lens_val]
                for lens in lensen:
                    ls[(o, lens)].append(sc[i])
    result = {}
    for o in RUBRIC_ONDERDELEN:
        for l in [1,2,3,4]:
            key = (o, l)
            inh = RUBRIC_INHOUD.get(key)
            if inh is None: result[key] = None; continue
            sl = ls[key]
            result[key] = round(sum(sl)/len(sl), 2) if sl else None
    return result

# ═══════════════════════════════════════════════════════════════
#  EXCEL EXPORT STUDENTEN
# ═══════════════════════════════════════════════════════════════
def excel_studenten(resultaten):
    DB="0F3460"; LB="EEF2FF"; WIT="FFFFFF"
    NH={1:"E74C3C",2:"E67E22",3:"F1C40F",4:"2ECC71",5:"1ABC9C"}
    thin=Side(style="thin",color="CCCCCC"); dik=Side(style="medium",color="0F3460")
    rand=Border(left=thin,right=thin,top=thin,bottom=thin)
    dik_rand=Border(left=dik,right=dik,top=dik,bottom=dik)
    wb=Workbook()
    secties=[s for s in VRAGEN_ST if s!="Overig"]

    ws=wb.active; ws.title="Samenvatting"; ws.sheet_view.showGridLines=False
    ws.column_dimensions["A"].width=26
    for c in ["B","C","D","E"]: ws.column_dimensions[c].width=18
    ws.merge_cells("A1:E1"); ws["A1"]="Cursusevaluatie Studenten - Resultaten Overzicht"
    ws["A1"].font=Font(name="Arial",bold=True,size=16,color=WIT)
    ws["A1"].fill=PatternFill("solid",fgColor=DB)
    ws["A1"].alignment=Alignment(horizontal="center",vertical="center"); ws.row_dimensions[1].height=36
    ws.merge_cells("A2:E2")
    ws["A2"]=f"Gegenereerd: {datetime.now().strftime('%d-%m-%Y %H:%M')}  |  Totaal: {len(resultaten)}"
    ws["A2"].font=Font(name="Arial",size=10,color="666666"); ws["A2"].fill=PatternFill("solid",fgColor=LB)
    ws["A2"].alignment=Alignment(horizontal="center",vertical="center"); ws.row_dimensions[2].height=20; ws.row_dimensions[3].height=8
    for ci,h in enumerate(["Sectie","Gem. score","Niveau (nr.)","Niveau (label)","Responsen"],1):
        c=ws.cell(row=4,column=ci,value=h); c.font=Font(name="Arial",bold=True,size=10,color=WIT)
        c.fill=PatternFill("solid",fgColor=DB)
        c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); c.border=rand
    ws.row_dimensions[4].height=28
    for rij,sectie in enumerate(secties,5):
        gems=[r["sectie_gemiddeldes"].get(sectie,0) for r in resultaten]
        gem=round(sum(gems)/len(gems),2) if gems else 0; niv=bereken_niveau(gem); kf=PatternFill("solid",fgColor=NH[niv])
        for ci,val in enumerate([sectie,gem,niv,NIVEAU_LABELS[niv],len(resultaten)],1):
            c=ws.cell(row=rij,column=ci,value=val); c.font=Font(name="Arial",size=10)
            c.alignment=Alignment(horizontal="center",vertical="center"); c.border=rand
            if ci in (3,4): c.fill=kf; c.font=Font(name="Arial",size=10,bold=True,color=WIT if niv in (1,2,5) else "1a1a2e")
        ws.row_dimensions[rij].height=22
    ls=5+len(secties)+1; ws.row_dimensions[ls-1].height=8
    ws.merge_cells(f"A{ls}:E{ls}"); ws[f"A{ls}"]="Niveauschaal"
    ws[f"A{ls}"].font=Font(name="Arial",bold=True,size=10,color=DB); ws[f"A{ls}"].fill=PatternFill("solid",fgColor=LB)
    ws[f"A{ls}"].alignment=Alignment(horizontal="left",vertical="center"); ws[f"A{ls}"].border=rand; ws.row_dimensions[ls].height=20
    for ro,(nr,lbl,bereik) in enumerate([(1,"Heel slecht","< 1.75"),(2,"Slecht","1.75-2.49"),(3,"Voldoende","2.50-3.49"),(4,"Goed","3.50-4.49"),(5,"Excellent",">= 4.50")],ls+1):
        kf=PatternFill("solid",fgColor=NH[nr])
        ws.merge_cells(f"A{ro}:B{ro}"); ws[f"A{ro}"]=f"Niveau {nr} - {lbl}"; ws[f"A{ro}"].fill=kf
        ws[f"A{ro}"].font=Font(name="Arial",size=9,bold=True,color=WIT if nr in (1,2,5) else "1a1a2e")
        ws[f"A{ro}"].alignment=Alignment(horizontal="center",vertical="center"); ws[f"A{ro}"].border=rand
        ws.merge_cells(f"C{ro}:E{ro}"); ws[f"C{ro}"]=f"Gemiddelde: {bereik}"
        ws[f"C{ro}"].font=Font(name="Arial",size=9); ws[f"C{ro}"].alignment=Alignment(horizontal="left",vertical="center"); ws[f"C{ro}"].border=rand; ws.row_dimensions[ro].height=18

    an=[1,2,3,4,5]; kp=[NIVEAU_KLEUREN[n] for n in an]
    df_t=pd.DataFrame([{f"N_{s.replace(' ','_')}":r["sectie_niveaus"].get(s,3) for s in secties} for r in resultaten])
    fig,axes=plt.subplots(2,3,figsize=(16,8)); fig.patch.set_facecolor("white"); af=axes.flatten()
    for idx,sectie in enumerate(secties):
        ax=af[idx]; col=f"N_{sectie.replace(' ','_')}"
        counts=df_t[col].value_counts().sort_index() if col in df_t.columns else pd.Series()
        cl=[counts.get(n,0) for n in an]
        bars=ax.bar(an,cl,color=kp,width=0.6,edgecolor="white",linewidth=1.5)
        for bar,cnt in zip(bars,cl):
            if cnt>0: ax.text(bar.get_x()+bar.get_width()/2,bar.get_height()+0.04,str(cnt),ha="center",va="bottom",fontsize=10,fontweight="bold",color="#1a1a2e")
        ax.set_title(sectie,fontsize=11,fontweight="bold",color="#0f3460",pad=6); ax.set_facecolor("white")
        ax.spines[["top","right"]].set_visible(False); ax.spines[["left","bottom"]].set_color("#ddd")
        ax.grid(axis="y",linestyle="--",alpha=0.3,color="#ccc"); ax.yaxis.set_major_locator(plt.MaxNLocator(integer=True)); ax.set_xticks(an)
    for idx in range(len(secties),len(af)): af[idx].set_visible(False)
    af[0].set_ylabel("Aantal",fontsize=9,color="#555"); af[3].set_ylabel("Aantal",fontsize=9,color="#555")
    fig.suptitle("Niveauverdeling per Sectie",fontsize=13,fontweight="bold",color="#0f3460",y=1.01); plt.tight_layout()
    ib=io.BytesIO(); fig.savefig(ib,format="png",dpi=130,bbox_inches="tight"); plt.close(fig); ib.seek(0)
    gr=ls+7; xl=XLImage(ib); xl.anchor=f"A{gr}"; xl.width=700; xl.height=440; ws.add_image(xl)

    ws2=wb.create_sheet("Gemiddelden per vraag"); ws2.sheet_view.showGridLines=False
    ws2.column_dimensions["A"].width=6; ws2.column_dimensions["B"].width=55; ws2.column_dimensions["C"].width=16; ws2.column_dimensions["D"].width=16
    ws2.merge_cells("A1:D1"); ws2["A1"]="Gemiddelde score per vraag"
    ws2["A1"].font=Font(name="Arial",bold=True,size=14,color=WIT); ws2["A1"].fill=PatternFill("solid",fgColor=DB)
    ws2["A1"].alignment=Alignment(horizontal="center",vertical="center"); ws2.row_dimensions[1].height=30
    rij=2
    for sectie,vl in VRAGEN_ST.items():
        if sectie=="Overig": continue
        ws2.row_dimensions[rij].height=8; rij+=1
        ws2.merge_cells(f"A{rij}:D{rij}"); ws2[f"A{rij}"]=sectie
        ws2[f"A{rij}"].font=Font(name="Arial",bold=True,size=11,color=WIT); ws2[f"A{rij}"].fill=PatternFill("solid",fgColor="1e3a5f")
        ws2[f"A{rij}"].alignment=Alignment(horizontal="left",vertical="center"); ws2[f"A{rij}"].border=rand; ws2.row_dimensions[rij].height=22; rij+=1
        for hi,h in enumerate(["#","Vraag","Gem. score","Lens"],1):
            c=ws2.cell(row=rij,column=hi,value=h); c.font=Font(name="Arial",bold=True,size=9,color=WIT)
            c.fill=PatternFill("solid",fgColor=DB); c.alignment=Alignment(horizontal="center",vertical="center"); c.border=rand
        ws2.row_dimensions[rij].height=18; rij+=1
        for vi,(vraag,lens) in enumerate(vl,1):
            av=[r.get("scores_per_vraag",{}).get(sectie,[])[vi-1] for r in resultaten if vi-1<len(r.get("scores_per_vraag",{}).get(sectie,[]))]
            gv=round(sum(av)/len(av),2) if av else "-"
            kfv=PatternFill("solid",fgColor=rubric_hex(gv if isinstance(gv,float) else None))
            ls_str="?" if (lens is None or lens=="open") else (f"Lens {','.join(map(str,lens))}" if isinstance(lens,list) else f"Lens {lens}")
            for ci,val in enumerate([vi,vraag,gv,ls_str],1):
                c=ws2.cell(row=rij,column=ci,value=val); c.font=Font(name="Arial",size=9)
                c.alignment=Alignment(horizontal="center" if ci!=2 else "left",vertical="center",wrap_text=(ci==2)); c.border=rand
                if ci==3 and isinstance(gv,float): c.fill=kfv
            ws2.row_dimensions[rij].height=30; rij+=1

    ws3=wb.create_sheet("Rubric"); ws3.sheet_view.showGridLines=False
    for ci,b in enumerate([28,36,36,36,36],1): ws3.column_dimensions[get_column_letter(ci)].width=b
    ws3.merge_cells("A1:E1"); ws3["A1"]="Rubric Analysemodel - Cursusevaluatie"
    ws3["A1"].font=Font(name="Arial",bold=True,size=14,color=WIT); ws3["A1"].fill=PatternFill("solid",fgColor=DB)
    ws3["A1"].alignment=Alignment(horizontal="center",vertical="center"); ws3.row_dimensions[1].height=34
    ws3.merge_cells("A2:E2"); ws3["A2"]="Kleur: Rood < 2.0  |  Oranje < 3.0  |  Geel < 4.0  |  Lichtgroen < 4.5  |  Groen >= 4.5  |  Grijs = geen data"
    ws3["A2"].font=Font(name="Arial",size=9,italic=True,color="444444"); ws3["A2"].fill=PatternFill("solid",fgColor=LB)
    ws3["A2"].alignment=Alignment(horizontal="center",vertical="center"); ws3.row_dimensions[2].height=18; ws3.row_dimensions[3].height=6
    for ci,h in enumerate(["Curriculumcomponent","Lens 1: Niveau & Samenhang\n(Biggs, Bloom, Dublin)","Lens 2: Didactisch ontwerp\n(Context-Concept, Entwistle, Gagne, Scaffolding)","Lens 3: Transfer theorie-praktijk\n(Kolb, Miller, Shulman PCK)","Lens 4: De student en leertools\n(TPACK, Zimmerman)"],1):
        c=ws3.cell(row=4,column=ci,value=h); c.font=Font(name="Arial",bold=True,size=9,color=WIT)
        c.fill=PatternFill("solid",fgColor="1e3a5f"); c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); c.border=rand
    ws3.row_dimensions[4].height=44
    lg=bereken_lens_gemiddeldes(resultaten) if resultaten else {}
    for ri,o in enumerate(RUBRIC_ONDERDELEN,5):
        ws3.row_dimensions[ri].height=110
        c=ws3.cell(row=ri,column=1,value=o); c.font=Font(name="Arial",bold=True,size=10)
        c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); c.border=dik_rand; c.fill=PatternFill("solid",fgColor=LB)
        for lens in [1,2,3,4]:
            gem=lg.get((o,lens)); inh=RUBRIC_INHOUD.get((o,lens)); tekst=inh if inh else "- (geen koppeling)"
            gs=f"Gem: {gem:.2f}\n\n" if (resultaten and gem is not None) else ""
            c=ws3.cell(row=ri,column=lens+1,value=gs+tekst); c.font=Font(name="Arial",size=8)
            c.alignment=Alignment(horizontal="left",vertical="top",wrap_text=True); c.border=rand; c.fill=PatternFill("solid",fgColor=rubric_hex(gem))

    ws4=wb.create_sheet("Open antwoorden"); ws4.sheet_view.showGridLines=False
    ws4.column_dimensions["A"].width=6; ws4.column_dimensions["B"].width=18; ws4.column_dimensions["C"].width=70
    ws4.merge_cells("A1:C1"); ws4["A1"]="Open antwoorden - Overig"
    ws4["A1"].font=Font(name="Arial",bold=True,size=14,color=WIT); ws4["A1"].fill=PatternFill("solid",fgColor=DB)
    ws4["A1"].alignment=Alignment(horizontal="center",vertical="center"); ws4.row_dimensions[1].height=30
    open_lijst=[r for r in resultaten if r.get("open_antwoord","").strip()]
    ws4.merge_cells("A2:C2"); ws4["A2"]=f"Ingevuld door {len(open_lijst)} van de {len(resultaten)} studenten"
    ws4["A2"].font=Font(name="Arial",size=10,color="666666"); ws4["A2"].fill=PatternFill("solid",fgColor=LB)
    ws4["A2"].alignment=Alignment(horizontal="center",vertical="center"); ws4.row_dimensions[2].height=18; ws4.row_dimensions[3].height=6
    for ci,h in enumerate(["#","Tijdstip","Antwoord"],1):
        c=ws4.cell(row=4,column=ci,value=h); c.font=Font(name="Arial",bold=True,size=10,color=WIT)
        c.fill=PatternFill("solid",fgColor=DB); c.alignment=Alignment(horizontal="center",vertical="center"); c.border=rand
    ws4.row_dimensions[4].height=22
    if open_lijst:
        for ri,r in enumerate(open_lijst,5):
            achter="F8F9FF" if ri%2==0 else "FFFFFF"
            for ci,val in enumerate([ri-4,r.get("tijdstip",""),r.get("open_antwoord","")],1):
                c=ws4.cell(row=ri,column=ci,value=val); c.font=Font(name="Arial",size=10)
                c.alignment=Alignment(horizontal="center" if ci!=3 else "left",vertical="top",wrap_text=True); c.border=rand; c.fill=PatternFill("solid",fgColor=achter)
            ws4.row_dimensions[ri].height=45
    else:
        ws4.merge_cells("A5:C5"); ws4["A5"]="Geen open antwoorden ingestuurd."
        ws4["A5"].font=Font(name="Arial",size=10,italic=True,color="888888"); ws4["A5"].alignment=Alignment(horizontal="center",vertical="center")

    ws5=wb.create_sheet("Cursus gehaald"); ws5.sheet_view.showGridLines=False
    ws5.column_dimensions["A"].width=6; ws5.column_dimensions["B"].width=22; ws5.column_dimensions["C"].width=36
    ws5.merge_cells("A1:C1"); ws5["A1"]="Cursus gehaald - Overzicht"
    ws5["A1"].font=Font(name="Arial",bold=True,size=14,color=WIT); ws5["A1"].fill=PatternFill("solid",fgColor=DB)
    ws5["A1"].alignment=Alignment(horizontal="center",vertical="center"); ws5.row_dimensions[1].height=30
    cg_opties=["Ja","Nee","Zeg ik liever niet / NVT"]
    cg_kleuren={"Ja":"D4EDDA","Nee":"F8D7DA","Zeg ik liever niet / NVT":"EEF2FF"}
    cg_counts={o:sum(1 for r in resultaten if r.get("cursus_gehaald","Zeg ik liever niet / NVT")==o) for o in cg_opties}
    totaal_cg=len(resultaten)
    ws5.merge_cells("A2:C2"); ws5["A2"]=f"Totaal responsen: {totaal_cg}"
    ws5["A2"].font=Font(name="Arial",size=10,color="666666"); ws5["A2"].fill=PatternFill("solid",fgColor=LB)
    ws5["A2"].alignment=Alignment(horizontal="center",vertical="center"); ws5.row_dimensions[2].height=18; ws5.row_dimensions[3].height=6
    for ci,h in enumerate(["#","Antwoord","Aantal (%)"],1):
        c=ws5.cell(row=4,column=ci,value=h); c.font=Font(name="Arial",bold=True,size=10,color=WIT)
        c.fill=PatternFill("solid",fgColor=DB); c.alignment=Alignment(horizontal="center",vertical="center"); c.border=rand
    ws5.row_dimensions[4].height=22
    for ri,(opt) in enumerate(cg_opties,5):
        cnt=cg_counts[opt]; pct=round(cnt/totaal_cg*100) if totaal_cg>0 else 0
        kf=PatternFill("solid",fgColor=cg_kleuren[opt])
        for ci,val in enumerate([ri-4,opt,f"{cnt}  ({pct}%)"],1):
            c=ws5.cell(row=ri,column=ci,value=val); c.font=Font(name="Arial",size=10)
            c.alignment=Alignment(horizontal="center",vertical="center"); c.border=rand; c.fill=kf
        ws5.row_dimensions[ri].height=22

    buf=io.BytesIO(); wb.save(buf); buf.seek(0); return buf.getvalue()

# ═══════════════════════════════════════════════════════════════
#  EXCEL EXPORT WERKVELD
# ═══════════════════════════════════════════════════════════════
def excel_werkveld(resultaten):
    DB="0F3460"; LB="EEF2FF"; GROEN="1E8449"; ROOD="922B21"; WIT="FFFFFF"
    NH={1:"E74C3C",2:"E67E22",3:"F1C40F",4:"2ECC71",5:"1ABC9C"}
    thin=Side(style="thin",color="DDDDDD"); rand=Border(left=thin,right=thin,top=thin,bottom=thin)
    wb=Workbook(); oks=list(VRAGEN_WV.keys())

    ws=wb.active; ws.title="Samenvatting"; ws.sheet_view.showGridLines=False
    ws.column_dimensions["A"].width=30
    for c in ["B","C","D","E"]: ws.column_dimensions[c].width=18
    ws.merge_cells("A1:E1"); ws["A1"]="Werkveld Cursusevaluatie - Resultaten Overzicht"
    ws["A1"].font=Font(name="Arial",bold=True,size=16,color=WIT); ws["A1"].fill=PatternFill("solid",fgColor=DB)
    ws["A1"].alignment=Alignment(horizontal="center",vertical="center"); ws.row_dimensions[1].height=36
    ws.merge_cells("A2:E2"); ws["A2"]=f"Gegenereerd: {datetime.now().strftime('%d-%m-%Y %H:%M')}  |  Totaal: {len(resultaten)}"
    ws["A2"].font=Font(name="Arial",size=10,color="666666"); ws["A2"].fill=PatternFill("solid",fgColor=LB)
    ws["A2"].alignment=Alignment(horizontal="center",vertical="center"); ws.row_dimensions[2].height=20; ws.row_dimensions[3].height=8
    for ci,h in enumerate(["Onderdeel","Gem. score","Niveau (nr.)","Niveau (label)","Responsen"],1):
        c=ws.cell(row=4,column=ci,value=h); c.font=Font(name="Arial",bold=True,size=10,color=WIT)
        c.fill=PatternFill("solid",fgColor=DB); c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); c.border=rand
    ws.row_dimensions[4].height=28
    for rij,key in enumerate(oks,5):
        sp=[r["scores"].get(key,0) for r in resultaten]
        gem=round(sum(sp)/len(sp),2) if sp else 0; niv=bereken_niveau(gem); kf=PatternFill("solid",fgColor=NH[niv])
        for ci,val in enumerate([key,gem,niv,NIVEAU_LABELS[niv],len(resultaten)],1):
            c=ws.cell(row=rij,column=ci,value=val); c.font=Font(name="Arial",size=10)
            c.alignment=Alignment(horizontal="center",vertical="center"); c.border=rand
            if ci in (3,4): c.fill=kf; c.font=Font(name="Arial",size=10,bold=True,color=WIT if niv in (1,2,5) else "1a1a2e")
        ws.row_dimensions[rij].height=22
    ls2=5+len(oks)+1; ws.row_dimensions[ls2-1].height=8
    ws.merge_cells(f"A{ls2}:E{ls2}"); ws[f"A{ls2}"]="Niveauschaal"
    ws[f"A{ls2}"].font=Font(name="Arial",bold=True,size=10,color=DB); ws[f"A{ls2}"].fill=PatternFill("solid",fgColor=LB)
    ws[f"A{ls2}"].alignment=Alignment(horizontal="left",vertical="center"); ws[f"A{ls2}"].border=rand; ws.row_dimensions[ls2].height=20
    for ro,(nr,lbl,bereik) in enumerate([(1,"Heel slecht","< 1.75"),(2,"Slecht","1.75-2.49"),(3,"Voldoende","2.50-3.49"),(4,"Goed","3.50-4.49"),(5,"Excellent",">= 4.50")],ls2+1):
        kf=PatternFill("solid",fgColor=NH[nr])
        ws.merge_cells(f"A{ro}:B{ro}"); ws[f"A{ro}"]=f"Niveau {nr} - {lbl}"; ws[f"A{ro}"].fill=kf
        ws[f"A{ro}"].font=Font(name="Arial",size=9,bold=True,color=WIT if nr in (1,2,5) else "1a1a2e")
        ws[f"A{ro}"].alignment=Alignment(horizontal="center",vertical="center"); ws[f"A{ro}"].border=rand
        ws.merge_cells(f"C{ro}:E{ro}"); ws[f"C{ro}"]=f"Gemiddelde: {bereik}"
        ws[f"C{ro}"].font=Font(name="Arial",size=9); ws[f"C{ro}"].alignment=Alignment(horizontal="left",vertical="center"); ws[f"C{ro}"].border=rand; ws.row_dimensions[ro].height=18

    wsg=wb.create_sheet("Gemiddelden per vraag"); wsg.sheet_view.showGridLines=False
    wsg.column_dimensions["A"].width=6; wsg.column_dimensions["B"].width=55; wsg.column_dimensions["C"].width=16
    wsg.merge_cells("A1:C1"); wsg["A1"]="Gemiddelde score per vraag (werkveld)"
    wsg["A1"].font=Font(name="Arial",bold=True,size=14,color=WIT); wsg["A1"].fill=PatternFill("solid",fgColor=DB)
    wsg["A1"].alignment=Alignment(horizontal="center",vertical="center"); wsg.row_dimensions[1].height=30
    rij=2
    for sectie,vl in VRAGEN_WV.items():
        wsg.row_dimensions[rij].height=8; rij+=1
        wsg.merge_cells(f"A{rij}:C{rij}"); wsg[f"A{rij}"]=sectie
        wsg[f"A{rij}"].font=Font(name="Arial",bold=True,size=11,color=WIT); wsg[f"A{rij}"].fill=PatternFill("solid",fgColor="1e3a5f")
        wsg[f"A{rij}"].alignment=Alignment(horizontal="left",vertical="center"); wsg[f"A{rij}"].border=rand; wsg.row_dimensions[rij].height=22; rij+=1
        for hi,h in enumerate(["#","Vraag","Gem. score"],1):
            c=wsg.cell(row=rij,column=hi,value=h); c.font=Font(name="Arial",bold=True,size=9,color=WIT)
            c.fill=PatternFill("solid",fgColor=DB); c.alignment=Alignment(horizontal="center",vertical="center"); c.border=rand
        wsg.row_dimensions[rij].height=18; rij+=1
        for vi,vraag in enumerate(vl,1):
            av=[r.get("scores_per_vraag",{}).get(sectie,[])[vi-1] for r in resultaten if vi-1<len(r.get("scores_per_vraag",{}).get(sectie,[]))]
            gv=round(sum(av)/len(av),2) if av else "-"
            kfv=PatternFill("solid",fgColor=rubric_hex(gv if isinstance(gv,float) else None))
            for ci,val in enumerate([vi,vraag,gv],1):
                c=wsg.cell(row=rij,column=ci,value=val); c.font=Font(name="Arial",size=9)
                c.alignment=Alignment(horizontal="center" if ci!=2 else "left",vertical="center",wrap_text=(ci==2)); c.border=rand
                if ci==3 and isinstance(gv,float): c.fill=kfv
            wsg.row_dimensions[rij].height=30; rij+=1

    ws3=wb.create_sheet("Ruwe data"); ws3.sheet_view.showGridLines=False
    rh=["#","Tijdstip","E-mailadres","Totaal niveau"]+[f"Gem. {k.split('-')[0].strip()}" for k in oks]+["Focusgroep"]
    cb=[5,18,30,14]+[16]*len(oks)+[14]
    for ci,(h,b) in enumerate(zip(rh,cb),1):
        ws3.column_dimensions[get_column_letter(ci)].width=b
        c=ws3.cell(row=1,column=ci,value=h); c.font=Font(name="Arial",bold=True,size=10,color=WIT)
        c.fill=PatternFill("solid",fgColor=DB); c.alignment=Alignment(horizontal="center",vertical="center"); c.border=rand
    ws3.row_dimensions[1].height=24
    for ri,r in enumerate(resultaten,2):
        tn=r["totaal_niveau"]; fg=r.get("focusgroep",False)
        kft=PatternFill("solid",fgColor=NH[tn]); fgf=PatternFill("solid",fgColor=GROEN if fg else ROOD)
        rv=[ri-1,r["tijdstip"],r.get("email",""),tn]+[r["scores"].get(k,"") for k in oks]+["Ja" if fg else "Nee"]
        for ci,val in enumerate(rv,1):
            c=ws3.cell(row=ri,column=ci,value=val); c.font=Font(name="Arial",size=9)
            c.alignment=Alignment(horizontal="center",vertical="center"); c.border=rand
            if ci==4: c.fill=kft; c.font=Font(name="Arial",size=9,bold=True,color=WIT if tn in (1,2,5) else "1a1a2e")
            if ci==len(rv): c.fill=fgf; c.font=Font(name="Arial",size=9,bold=True,color=WIT)
        ws3.row_dimensions[ri].height=18

    ws4=wb.create_sheet("Focusgroep aanmeldingen"); ws4.sheet_view.showGridLines=False
    ws4.column_dimensions["A"].width=5; ws4.column_dimensions["B"].width=35; ws4.column_dimensions["C"].width=20
    ws4.merge_cells("A1:C1"); ws4["A1"]="Focusgroep - Geinteresseerde deelnemers"
    ws4["A1"].font=Font(name="Arial",bold=True,size=14,color=WIT); ws4["A1"].fill=PatternFill("solid",fgColor=DB)
    ws4["A1"].alignment=Alignment(horizontal="center",vertical="center"); ws4.row_dimensions[1].height=32
    fg_lijst=[r for r in resultaten if r.get("focusgroep")]
    ws4.merge_cells("A2:C2"); ws4["A2"]=f"Totaal aangemeld: {len(fg_lijst)}"
    ws4["A2"].font=Font(name="Arial",size=10,color="444444"); ws4["A2"].fill=PatternFill("solid",fgColor=LB)
    ws4["A2"].alignment=Alignment(horizontal="center",vertical="center"); ws4.row_dimensions[2].height=20; ws4.row_dimensions[3].height=8
    for ci,h in enumerate(["#","E-mailadres","Tijdstip"],1):
        c=ws4.cell(row=4,column=ci,value=h); c.font=Font(name="Arial",bold=True,size=10,color=WIT)
        c.fill=PatternFill("solid",fgColor=GROEN); c.alignment=Alignment(horizontal="center",vertical="center"); c.border=rand
    ws4.row_dimensions[4].height=24
    if fg_lijst:
        for ri,r in enumerate(fg_lijst,5):
            for ci,val in enumerate([ri-4,r.get("email",""),r["tijdstip"]],1):
                c=ws4.cell(row=ri,column=ci,value=val); c.font=Font(name="Arial",size=10)
                c.alignment=Alignment(horizontal="center",vertical="center"); c.border=rand
                if ri%2==0: c.fill=PatternFill("solid",fgColor="F0FFF4")
            ws4.row_dimensions[ri].height=20
    else:
        ws4.merge_cells("A5:C5"); ws4["A5"]="Geen aanmeldingen."
        ws4["A5"].font=Font(name="Arial",size=10,italic=True,color="888888"); ws4["A5"].alignment=Alignment(horizontal="center",vertical="center")

    buf=io.BytesIO(); wb.save(buf); buf.seek(0); return buf.getvalue()

# ═══════════════════════════════════════════════════════════════
#  STIJL
# ═══════════════════════════════════════════════════════════════
def laad_stijl():
    st.markdown("""
    <style>
        @import url('https://fonts.googleapis.com/css2?family=DM+Serif+Display&family=DM+Sans:wght@300;400;500;600&display=swap');
        html,body,[class*="css"]{font-family:'DM Sans',sans-serif;}
        h1,h2,h3{font-family:'DM Serif Display',serif;}
        .landing{background:linear-gradient(135deg,#1a1a2e 0%,#16213e 50%,#0f3460 100%);
            color:white;padding:3.5rem 2rem 3rem;border-radius:20px;margin-bottom:2rem;
            text-align:center;box-shadow:0 10px 40px rgba(15,52,96,0.5);}
        .landing h1{font-size:2.8rem;margin:0;letter-spacing:-0.5px;}
        .landing p{opacity:0.75;margin:0.8rem 0 2rem;font-size:1.1rem;}
        .main-header{background:linear-gradient(135deg,#1a1a2e 0%,#16213e 50%,#0f3460 100%);
            color:white;padding:2.5rem 2rem 2rem;border-radius:16px;margin-bottom:1.5rem;
            text-align:center;box-shadow:0 8px 32px rgba(15,52,96,0.4);}
        .main-header h1{font-size:2.2rem;margin:0;letter-spacing:-0.5px;}
        .main-header p{opacity:0.80;margin:0.6rem 0 0;font-size:0.97rem;line-height:1.6;}
        .schaal-info{background:#fffbeb;border:1.5px solid #fcd34d;border-radius:12px;padding:1.2rem 1.6rem;margin-bottom:1.5rem;}
        .schaal-info h4{color:#92400e;margin:0 0 0.6rem;font-size:1rem;}
        .schaal-info p,.schaal-info li{color:#78350f;font-size:0.88rem;margin:0.2rem 0;}
        .sh-card{background:#f0f9ff;border:2px solid #7dd3fc;border-radius:14px;padding:1.4rem 1.8rem;margin-bottom:1.5rem;}
        .sh-card h3{color:#0c4a6e;margin-top:0;}
        .email-card{background:#f8f9ff;border:2px solid #c7d2fe;border-radius:14px;padding:2rem;margin-bottom:1.5rem;text-align:center;}
        .email-card h3{color:#0f3460;margin-top:0;}
        .sectie-card{background:#f8f9ff;border-left:5px solid #0f3460;border-radius:10px;padding:1.2rem 1.6rem;margin-bottom:1.2rem;}
        .sectie-card h3{color:#0f3460;margin-top:0;font-size:1.1rem;}
        .focusgroep-card{background:#f0fdf4;border:2px solid #86efac;border-radius:14px;padding:1.6rem;margin-bottom:1.5rem;}
        .focusgroep-card h3{color:#166534;margin-top:0;}
        .rubric-card{border-radius:12px;padding:1.4rem;margin-bottom:0.8rem;color:white;box-shadow:0 4px 16px rgba(0,0,0,0.12);}
        .rubric-card h4{margin:0 0 0.3rem;font-size:0.95rem;opacity:0.85;}
        .rubric-card .badge{font-family:'DM Serif Display',serif;font-size:1.4rem;font-weight:bold;}
        .rubric-card .sub{margin-top:0.4rem;font-size:0.85rem;opacity:0.9;}
        .totaal-badge{text-align:center;border-radius:14px;padding:1.8rem;margin-top:1.5rem;color:white;box-shadow:0 6px 24px rgba(0,0,0,0.2);}
        .totaal-badge .label{font-size:0.9rem;opacity:0.8;text-transform:uppercase;letter-spacing:1px;}
        .totaal-badge .tekst{font-family:'DM Serif Display',serif;font-size:2rem;margin-top:0.2rem;}
        .fg-ja{background:#dcfce7;border:1.5px solid #86efac;border-radius:10px;padding:1rem 1.4rem;color:#166534;font-weight:600;}
        .fg-nee{background:#fef2f2;border:1.5px solid #fca5a5;border-radius:10px;padding:1rem 1.4rem;color:#991b1b;font-weight:600;}
        .stButton>button{background:#0f3460;color:white;border:none;border-radius:8px;padding:0.6rem 2rem;font-family:'DM Sans',sans-serif;font-size:1rem;font-weight:600;transition:background 0.2s;}
        .stButton>button:hover{background:#e94560;}
        .dashboard-header{background:linear-gradient(135deg,#0f3460,#e94560);color:white;padding:1.5rem 2rem;border-radius:12px;margin-bottom:1.5rem;}
        .rubric-tabel th{background:#1e3a5f;color:white;text-align:center;padding:10px 8px;font-size:0.82rem;}
        .rubric-tabel td{border:1px solid #ddd;padding:10px;font-size:0.8rem;vertical-align:top;line-height:1.5;}
        .rubric-tabel tr td:first-child{font-weight:700;background:#eef2ff;color:#0f3460;text-align:center;font-size:0.9rem;}
    </style>""", unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════
#  LANDINGSPAGINA
# ═══════════════════════════════════════════════════════════════
def landingspagina():
    st.markdown("""
    <div class="landing">
        <h1>Cursusevaluatie ALO</h1>
        <p>Kies uw rol om de evaluatie te starten of de resultaten te bekijken.</p>
    </div>""", unsafe_allow_html=True)
    c1, c2, c3 = st.columns(3)
    with c1:
        st.markdown("#### Student")
        st.caption("Vul de anonieme cursusevaluatie in.")
        if st.button("Ik ben Student", use_container_width=True):
            st.session_state["rol"] = "student"; st.rerun()
    with c2:
        st.markdown("#### Stagebegeleider")
        st.caption("Beoordeel de cursus vanuit het werkveld.")
        if st.button("Ik ben Stagebegeleider", use_container_width=True):
            st.session_state["rol"] = "werkveld"; st.rerun()
    with c3:
        st.markdown("#### Docent")
        st.caption("Bekijk de evaluatieresultaten.")
        if st.button("Ik ben Docent", use_container_width=True):
            st.session_state["rol"] = "docent"; st.rerun()

# ═══════════════════════════════════════════════════════════════
#  STUDENT PAGINA
# ═══════════════════════════════════════════════════════════════
def student_pagina():
    if st.button("<- Terug naar startpagina"):
        st.session_state["rol"] = None; st.rerun()
    st.markdown("""
    <div class="main-header">
        <h1>Cursus evaluatie studenten</h1>
        <p>Beantwoord de vragen eerlijk - de evaluatie is volledig anoniem.<br>
        Vul niet het gewenste juiste antwoord in; dit leidt tot minder accurate resultaten.<br>
        <em>Er wordt gebruik gemaakt van de Likert-schaal (voor hulp klik op de (?) naast de vraag).</em></p>
    </div>""", unsafe_allow_html=True)
    st.markdown(SCHAAL_INFO, unsafe_allow_html=True)
    st.divider()
    alle_scores = {}
    open_antwoord_tekst = ""
    for onderdeel, vraaglijst in VRAGEN_ST.items():
        st.markdown(f'<div class="sectie-card"><h3>📌 {onderdeel}</h3></div>', unsafe_allow_html=True)
        if onderdeel == "Overig":
            open_antwoord_tekst = st.text_area(
                f"**1.** {vraaglijst[0][0]}",
                placeholder="Typ hier uw opmerkingen, tips of suggesties...",
                key="st_overig_open", height=120)
            st.markdown(f"**2.** Heb je de cursus gehaald?")
            cursus_gehaald_antw = st.radio(
                "Cursus gehaald",
                options=["Ja", "Nee", "Zeg ik liever niet / NVT"],
                index=2,
                horizontal=True,
                key="st_cursus_gehaald",
                label_visibility="collapsed")
            alle_scores[onderdeel] = []
        else:
            sc = []
            for i, (vraag, _) in enumerate(vraaglijst, 1):
                sc.append(st.slider(f"**{i}.** {vraag}", 1, 5, 3, key=f"st_{onderdeel}_{i}",
                    help="1 = sterk mee oneens  |  3 = neutraal  |  5 = sterk mee eens"))
            alle_scores[onderdeel] = sc
        st.markdown("---")
    if st.button("Stuur mijn antwoorden in"):
        sg, sn = {}, {}
        for o, sc in alle_scores.items():
            if not sc: continue
            g = sum(sc)/len(sc); sg[o] = round(g,2); sn[o] = bereken_niveau(g)
        tg = sum(sg.values())/len(sg); tn = bereken_niveau(tg)
        cursus_gehaald_antw = st.session_state.get("st_cursus_gehaald", "Zeg ik liever niet / NVT")
        sla_student_op(alle_scores, sg, sn, tn, open_antwoord_tekst, cursus_gehaald_antw)
        st.session_state["st_ingediend"] = True
        st.session_state["st_resultaat"] = {
            "niveaus": sn, "gemiddeldes": sg, "totaal_niveau": tn
        }
        st.rerun()


# ═══════════════════════════════════════════════════════════════
#  STUDENT - bedankt scherm
# ═══════════════════════════════════════════════════════════════
def st_bedankt():
    res = st.session_state.get("st_resultaat", {})
    niveaus = res.get("niveaus", {}); gemiddeldes = res.get("gemiddeldes", {})
    tn = res.get("totaal_niveau", 3)

    st.markdown("""
    <div class="main-header">
        <h1>Cursus evaluatie studenten</h1>
        <p>Bedankt voor het invullen van de evaluatie!</p>
    </div>""", unsafe_allow_html=True)

    st.success("✅ Je antwoorden zijn succesvol ingestuurd. Hartelijk dank!")
    st.markdown("## Jouw mening over de cursus")

    cols = st.columns(min(len(niveaus), 3))
    for idx, (o, niv) in enumerate(niveaus.items()):
        with cols[idx % 3]:
            st.markdown(f'<div class="rubric-card" style="background:{niveau_kleur_css(niv)};"><h4>{o}</h4><div class="badge">{NIVEAU_LABELS[niv]}</div><div class="sub">Gemiddelde: {gemiddeldes[o]:.2f}</div></div>', unsafe_allow_html=True)

    kt = niveau_kleur_css(tn)
    st.markdown(f'<div class="totaal-badge" style="background:linear-gradient(135deg,{kt}cc,{kt});"><div class="label">Totaal Niveau</div><div class="tekst">{NIVEAU_LABELS[tn]}</div><div style="margin-top:0.5rem;opacity:0.85;font-size:0.93rem;">{NIVEAU_BESCHRIJVING[tn]}</div></div>', unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════
#  WERKVELD - email scherm
# ═══════════════════════════════════════════════════════════════
def wv_email_scherm():
    if st.button("<- Terug naar startpagina"):
        st.session_state["rol"] = None; st.rerun()
    st.markdown("""
    <div class="main-header">
        <h1>Werkveld Cursusevaluatie</h1>
        <p>Bedankt voor uw deelname. Vul eerst uw e-mailadres in om te beginnen.<br>
        <em>Uw e-mailadres wordt alleen bewaard als u interesse heeft in de focusgroep.</em></p>
    </div>""", unsafe_allow_html=True)
    st.markdown("""<div class="email-card"><h3>Uw e-mailadres</h3>
    <p style="color:#555;margin-bottom:0;">Uw e-mailadres wordt alleen gebruikt voor eventuele
    follow-up via de focusgroep. De evaluatieresultaten zelf zijn niet aan u te koppelen.</p>
    </div>""", unsafe_allow_html=True)
    email = st.text_input("E-mailadres", placeholder="naam@organisatie.nl", label_visibility="collapsed")
    c1, c2, c3 = st.columns([1,2,1])
    with c2:
        if st.button("Doorgaan naar de evaluatie", use_container_width=True):
            if not email.strip(): st.error("Vul een e-mailadres in.")
            elif not is_geldig_email(email): st.error("Voer een geldig e-mailadres in.")
            else: st.session_state["wv_email"] = email.strip().lower(); st.rerun()

# ═══════════════════════════════════════════════════════════════
#  WERKVELD - studiehandleiding scherm
# ═══════════════════════════════════════════════════════════════
def wv_studiehandleiding_scherm():
    st.markdown("""
    <div class="main-header">
        <h1>Werkveld Cursusevaluatie</h1>
        <p>Voordat u begint: het is van waarde dat u de inhoud van de cursus kent.<br>
        Hieronder kunt u de studiehandleiding openen ter voorbereiding op de evaluatie.</p>
    </div>""", unsafe_allow_html=True)
    if os.path.exists(STUDIEHANDLEIDING_PAD):
        with open(STUDIEHANDLEIDING_PAD, "rb") as f:
            pdf_bytes = f.read()
        b64 = base64.b64encode(pdf_bytes).decode()
        st.markdown(f"""
        <div class="sh-card">
            <h3>Studiehandleiding - {STUDIEHANDLEIDING_NAAM}</h3>
            <p style="color:#0c4a6e;margin-bottom:1.2rem;">
                Klik op de knop hieronder om de studiehandleiding te openen in een nieuw tabblad.
                U kunt daarna terugkeren naar deze pagina om de evaluatie in te vullen. <br>
                **LET OP je moet het geopende pdf bestand refreshen mocht u de informatie niet zien.**
            </p>
            <a href="data:application/pdf;base64,{b64}" target="_blank"
               style="background:#0f3460;color:white;padding:0.6rem 1.4rem;border-radius:8px;
                      text-decoration:none;font-weight:600;font-size:0.95rem;">
               Open studiehandleiding (PDF)
            </a>
        </div>""", unsafe_allow_html=True)
    else:
        st.warning(f"Studiehandleiding niet gevonden. Zet het PDF-bestand op pad: {STUDIEHANDLEIDING_PAD}")
        st.info("Tip: Wijzig STUDIEHANDLEIDING_PAD bovenin het script naar de juiste bestandsnaam.")
    st.markdown("&nbsp;")
    c1, c2, c3 = st.columns([1,2,1])
    with c2:
        if st.button("Doorgaan naar de evaluatie", use_container_width=True):
            st.session_state["wv_sh_gezien"] = True; st.rerun()

# ═══════════════════════════════════════════════════════════════
#  WERKVELD - vragenlijst
# ═══════════════════════════════════════════════════════════════
def wv_vragenlijst():
    st.markdown("""
    <div class="main-header">
        <h1>Werkveld cursusevaluatie</h1>
        <p>Beantwoord de vragen eerlijk. De evaluatie is volledig anoniem mocht u geen focusgroep willen.<br>
        Deze vragenlijst heeft als doelstelling het verbeteren van de cursus voor de volgende generatie stagiaires.<br>
        <em>Er wordt gebruik gemaakt van de Likert-schaal (voor hulp klik op de (?) naast de vraag).</em></p>
    </div>""", unsafe_allow_html=True)
    st.markdown(SCHAAL_INFO, unsafe_allow_html=True)
    st.divider()
    alle_scores = {}
    alle_spv = {}
    for onderdeel, vragen_lijst in VRAGEN_WV.items():
        letter = onderdeel.split("-")[0].strip()
        st.markdown(f'<div class="sectie-card"><h3>📌 {onderdeel}</h3></div>', unsafe_allow_html=True)
        sc = []
        for i, vraag in enumerate(vragen_lijst, 1):
            sc.append(st.slider(f"**V{letter}{i}.** {vraag}", 1, 5, 3, key=f"wv_{onderdeel}_{i}",
                help="1 = sterk mee oneens  |  3 = neutraal  |  5 = sterk mee eens"))
        alle_scores[onderdeel] = sc; alle_spv[onderdeel] = sc
        st.markdown("---")
    st.markdown('<div class="focusgroep-card"><h3>C - Focusgroep</h3></div>', unsafe_allow_html=True)
    fg_antw = st.radio(
        "Heeft u behoefte aan een Teams meeting met docenten om nog verdere feedback "
        "of andere punten te bespreken over de cursus?",
        options=["Ja","Nee"], index=1, horizontal=True, key="wv_fg_radio")
    st.markdown("---")
    c1, c2, c3 = st.columns([1,2,1])
    with c2:
        if st.button("Stuur mijn evaluatie in", use_container_width=True):
            niveaus, gemiddeldes = {}, {}
            for o, sc in alle_scores.items():
                g = sum(sc)/len(sc); gemiddeldes[o] = round(g,2); niveaus[o] = bereken_niveau(g)
            tg = sum(gemiddeldes.values())/len(gemiddeldes); tn = bereken_niveau(tg)
            fg_ja = fg_antw == "Ja"
            sla_werkveld_op(st.session_state["wv_email"], gemiddeldes, alle_spv, niveaus, tn, fg_ja)
            st.session_state["wv_ingediend"] = True
            st.session_state["wv_resultaat"] = {"niveaus": niveaus, "gemiddeldes": gemiddeldes, "totaal_niveau": tn, "focusgroep": fg_ja}
            st.rerun()

# ═══════════════════════════════════════════════════════════════
#  WERKVELD - bedankt scherm
# ═══════════════════════════════════════════════════════════════
def wv_bedankt():
    res = st.session_state.get("wv_resultaat", {})
    niveaus = res.get("niveaus",{}); gemiddeldes = res.get("gemiddeldes",{})
    tn = res.get("totaal_niveau",3); fg = res.get("focusgroep",False)
    st.success("Uw evaluatie is succesvol ingediend. Hartelijk dank!")
    st.markdown("## Uw mening over de cursus")
    cols = st.columns(len(niveaus))
    for col,(o,niv) in zip(cols,niveaus.items()):
        with col:
            st.markdown(f'<div class="rubric-card" style="background:{niveau_kleur_css(niv)};"><h4>{o}</h4><div class="badge">{NIVEAU_LABELS[niv]}</div><div class="sub">Gemiddelde: {gemiddeldes[o]:.2f}</div></div>', unsafe_allow_html=True)
    kt = niveau_kleur_css(tn)
    st.markdown(f'<div class="totaal-badge" style="background:linear-gradient(135deg,{kt}cc,{kt});"><div class="label">Totaal Niveau</div><div class="tekst">{NIVEAU_LABELS[tn]}</div><div style="margin-top:0.5rem;opacity:0.85;font-size:0.93rem;">{NIVEAU_BESCHRIJVING[tn]}</div></div>', unsafe_allow_html=True)
    st.markdown("&nbsp;")
    if fg:
        st.markdown('<div class="fg-ja">U heeft zich aangemeld voor de focusgroep. Wij nemen contact op via uw e-mailadres.</div>', unsafe_allow_html=True)
    else:
        st.markdown('<div class="fg-nee">U heeft geen interesse in de focusgroep aangegeven. Bedankt voor uw deelname!</div>', unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════
#  DOCENT DASHBOARD - studenten
# ═══════════════════════════════════════════════════════════════
def dash_studenten():
    st.markdown("""
    <div class="dashboard-header">
        <h2 style="margin:0;font-family:'DM Serif Display',serif;">Dashboard - Studentenevaluatie | Voor docenten zelf evaluatie zie hierboven.</h2>
        <p style="margin:0.4rem 0 0;opacity:0.85;">Academie voor Lichamelijke Opvoeding - resultaten van studenten. <br>
        Graag na het downloaden van de resultaten de resultaten verwijderen zodat de volgende gerbuiker hier betrouwbare informatie uithaald.</p>
    </div>""", unsafe_allow_html=True)
    resultaten = laad(TABEL_ST)
    if not resultaten:
        st.warning("Er zijn nog geen studentenevaluaties ingestuurd."); return
    aantal = len(resultaten)
    st.markdown(f"""
    <div style="background:#eef2ff;border:1px solid #c7d2fe;border-radius:12px;
                padding:1.2rem 1.8rem;margin-bottom:1.5rem;display:flex;align-items:center;gap:1.2rem;">
        <div style="font-size:2.8rem;font-family:'DM Serif Display',serif;color:#0f3460;line-height:1;">{aantal}</div>
        <div><div style="font-weight:600;color:#0f3460;font-size:1rem;">{"respons ontvangen" if aantal==1 else "responsen ontvangen"}</div>
        <div style="color:#666;font-size:0.85rem;">Anonieme cursusevaluaties van studenten</div></div>
    </div>""", unsafe_allow_html=True)

    st.subheader("Niveauverdeling per sectie")
    an=[1,2,3,4,5]; kl=[NIVEAU_KLEUREN[n] for n in an]
    sn=[s for s in VRAGEN_ST if s!="Overig"]
    df_niv=pd.DataFrame([{s:r["sectie_niveaus"].get(s,3) for s in sn} for r in resultaten])
    fig,axes=plt.subplots(2,3,figsize=(16,8)); fig.patch.set_facecolor("#f8f9ff"); af=axes.flatten()
    for idx,sectie in enumerate(sn):
        ax=af[idx]; counts=df_niv[sectie].value_counts().sort_index() if sectie in df_niv.columns else pd.Series()
        cl=[counts.get(n,0) for n in an]
        bars=ax.bar(an,cl,color=kl,width=0.6,edgecolor="white",linewidth=1.5)
        for bar,cnt in zip(bars,cl):
            if cnt>0: ax.text(bar.get_x()+bar.get_width()/2,bar.get_height()+0.05,str(cnt),ha="center",va="bottom",fontsize=11,fontweight="bold",color="#1a1a2e")
        ax.set_title(sectie,fontsize=11,fontweight="bold",color="#0f3460",pad=8); ax.set_facecolor("#f8f9ff")
        ax.spines[["top","right"]].set_visible(False); ax.spines[["left","bottom"]].set_color("#ddd")
        ax.grid(axis="y",linestyle="--",alpha=0.3,color="#ccc"); ax.yaxis.set_major_locator(plt.MaxNLocator(integer=True)); ax.set_xticks(an)
    for idx in range(len(sn),len(af)): af[idx].set_visible(False)
    af[0].set_ylabel("Aantal inzendingen",fontsize=9,color="#555"); af[3].set_ylabel("Aantal inzendingen",fontsize=9,color="#555")
    fig.suptitle("Niveauverdeling per Sectie",fontsize=14,fontweight="bold",color="#0f3460",y=1.01)
    plt.tight_layout(); st.pyplot(fig); plt.close(fig)
    st.markdown("---")

    st.subheader("Gemiddelde scores per vraag & per sectie")
    for sectie, vl in VRAGEN_ST.items():
        if sectie == "Overig": continue
        with st.expander(f"📌 {sectie}", expanded=False):
            gl=[r["sectie_gemiddeldes"].get(sectie,0) for r in resultaten]
            gs=round(sum(gl)/len(gl),2) if gl else 0; ns=bereken_niveau(gs); ks=niveau_kleur_css(ns)
            st.markdown(f'<div style="background:{ks}22;border-left:4px solid {ks};border-radius:8px;padding:0.7rem 1rem;margin-bottom:1rem;"><strong>Sectiegemiddelde: {gs:.2f}</strong> &nbsp;->&nbsp; {NIVEAU_LABELS[ns]}</div>', unsafe_allow_html=True)
            rows=[]
            for vi,(vraag,lens) in enumerate(vl,1):
                av=[r.get("scores_per_vraag",{}).get(sectie,[])[vi-1] for r in resultaten if vi-1<len(r.get("scores_per_vraag",{}).get(sectie,[]))]
                gv=round(sum(av)/len(av),2) if av else "-"
                ls_str="?" if (lens is None or lens=="open") else (f"Lens {','.join(map(str,lens))}" if isinstance(lens,list) else f"Lens {lens}")
                rows.append({"Nr.":vi,"Vraag":vraag,"Gemiddelde":gv,"Lens":ls_str})
            st.dataframe(pd.DataFrame(rows),use_container_width=True,hide_index=True)
    st.markdown("---")

    st.subheader("Rubric - Analysemodel")
    st.caption("Voor een verduidelijking van de rubric zie het eindwerk van Jesper Visser.  De kleur per cel geeft het gemiddelde van de bijbehorende vraag aan.")
    lg=bereken_lens_gemiddeldes(resultaten)
    lh=["Lens 1: Niveau & Samenhang<br><small>(Biggs, Bloom, Dublin)</small>","Lens 2: Didactisch ontwerp<br><small>(Context-Concept, Entwistle, Gagne, Scaffolding)</small>","Lens 3: Transfer theorie-praktijk<br><small>(Kolb, Miller, Shulman PCK)</small>","Lens 4: De student en leertools<br><small>(TPACK, Zimmerman)</small>"]
    st.markdown("""<div style="display:flex;gap:1rem;flex-wrap:wrap;margin-bottom:0.8rem;font-size:0.8rem;">
    <span style="background:#f8d7da;padding:3px 10px;border-radius:6px;">Rood: Gem &lt; 2.0</span>
    <span style="background:#fde8c8;padding:3px 10px;border-radius:6px;">Oranje: Gem &lt; 3.0</span>
    <span style="background:#fff9c4;padding:3px 10px;border-radius:6px;">Geel: Gem &lt; 4.0</span>
    <span style="background:#d4edda;padding:3px 10px;border-radius:6px;">Lichtgroen: Gem &lt; 4.5</span>
    <span style="background:#a8e6cf;padding:3px 10px;border-radius:6px;">Groen: Gem >= 4.5</span>
    <span style="background:#e8ecf4;padding:3px 10px;border-radius:6px;">Grijs: Geen data</span></div>""", unsafe_allow_html=True)
    tabel='<table class="rubric-tabel" style="width:100%;border-collapse:collapse;"><thead><tr>'
    tabel+='<th style="background:#1e3a5f;color:white;padding:10px;width:14%;">Curriculumcomponent</th>'
    for h in lh: tabel+=f'<th style="background:#1e3a5f;color:white;padding:10px;width:21.5%;">{h}</th>'
    tabel+="</tr></thead><tbody>"
    for o in RUBRIC_ONDERDELEN:
        tabel+="<tr>"
        tabel+=f'<td style="background:#eef2ff;font-weight:700;color:#0f3460;text-align:center;padding:10px;border:1px solid #ddd;">{o}</td>'
        for lens in [1,2,3,4]:
            gem=lg.get((o,lens)); inh=RUBRIC_INHOUD.get((o,lens)); achter=rubric_kleur(gem)
            tekst=(inh or "- (geen koppeling)").replace("\n","<br>")
            gb=f'<div style="font-weight:700;font-size:0.78rem;color:{rubric_kleur_tekst(gem)};margin-bottom:6px;">Gemiddelde: {gem:.2f}</div>' if gem is not None else ""
            tabel+=f'<td style="background:{achter};border:1px solid #ccc;padding:10px;font-size:0.78rem;vertical-align:top;line-height:1.55;">{gb}{tekst}</td>'
        tabel+="</tr>"
    tabel+="</tbody></table>"
    st.markdown(tabel, unsafe_allow_html=True)
    st.markdown("---")

    st.subheader("Open antwoorden - Overig")
    open_antwoorden=[{"Tijdstip":r.get("tijdstip",""),"Antwoord":r.get("open_antwoord","")} for r in resultaten if r.get("open_antwoord","").strip()]
    if open_antwoorden:
        st.caption(f"{len(open_antwoorden)} van de {len(resultaten)} studenten heeft iets ingevuld.")
        for idx,item in enumerate(open_antwoorden,1):
            st.markdown(f'<div style="background:#f8f9ff;border-left:4px solid #0f3460;border-radius:8px;padding:0.8rem 1.2rem;margin-bottom:0.6rem;"><div style="font-size:0.78rem;color:#888;margin-bottom:0.3rem;">#{idx} - {item["Tijdstip"]}</div><div style="font-size:0.95rem;color:#1a1a2e;">{item["Antwoord"]}</div></div>', unsafe_allow_html=True)
    else:
        st.info("Er zijn nog geen open antwoorden ingestuurd.")
    st.markdown("---")

    st.subheader("Cursus gehaald - Overzicht")
    cg_counts = {"Ja": 0, "Nee": 0, "Zeg ik liever niet / NVT": 0}
    for r in resultaten:
        antw = r.get("cursus_gehaald", "Zeg ik liever niet / NVT")
        if antw in cg_counts:
            cg_counts[antw] += 1
        else:
            cg_counts["Zeg ik liever niet / NVT"] += 1
    totaal_cg = len(resultaten)
    c1, c2, c3 = st.columns(3)
    kleuren_cg = {"Ja": ("#d4edda","#155724"), "Nee": ("#f8d7da","#721c24"), "Zeg ik liever niet / NVT": ("#eef2ff","#0f3460")}
    for col, (optie, achter_tekst) in zip([c1,c2,c3], kleuren_cg.items()):
        achter, tekst = achter_tekst
        cnt = cg_counts[optie]
        pct = round(cnt/totaal_cg*100) if totaal_cg > 0 else 0
        with col:
            st.markdown(f'<div style="background:{achter};border-radius:10px;padding:1rem 1.2rem;text-align:center;"><div style="font-size:2rem;font-weight:700;color:{tekst};">{cnt}</div><div style="font-weight:600;color:{tekst};font-size:0.9rem;">{optie}</div><div style="color:#555;font-size:0.8rem;">{pct}% van de responsen</div></div>', unsafe_allow_html=True)
    st.markdown("---")

    st.subheader("Resultaten downloaden")
    st.download_button("Download als Excel (.xlsx)", excel_studenten(resultaten),
        f"studenten_evaluatie_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
    st.markdown("---")
    if st.button("Verwijder alle studentenresultaten"):
        verwijder_alle(TABEL_ST)
        st.success("Alle studentenresultaten zijn verwijderd."); st.rerun()

# ═══════════════════════════════════════════════════════════════
#  DOCENT DASHBOARD - werkveld
# ═══════════════════════════════════════════════════════════════
def dash_werkveld():
    st.markdown("""
    <div class="dashboard-header">
        <h2 style="margin:0;font-family:'DM Serif Display',serif;">Dashboard - Werkveld evaluatie</h2>
        <p style="margin:0.3rem 0 0;opacity:0.8;">Overzicht van alle ingestuurde werkveld evaluaties. <br> 
        Graag na het downloaden van de resultaten de resultaten verwijderen zodat de volgende gerbuiker hier betrouwbare informatie uithaald.</p>
    </div>""", unsafe_allow_html=True)
    resultaten = laad(TABEL_WV)
    if not resultaten:
        st.warning("Er zijn nog geen werkveld evaluaties ingestuurd."); return
    oks=list(VRAGEN_WV.keys())
    df=pd.DataFrame([{"Totaal niveau":r["totaal_niveau"],**{f"Niveau {k.split('-')[0].strip()}":r["niveaus"].get(k,3) for k in oks},"Focusgroep":r.get("focusgroep",False)} for r in resultaten])
    aantal=len(df); fg_aantal=int(df["Focusgroep"].sum())
    c1,c2=st.columns(2)
    with c1:
        st.markdown(f"""<div style="background:#eef2ff;border:1px solid #c7d2fe;border-radius:12px;
        padding:1.2rem 1.8rem;display:flex;align-items:center;gap:1.2rem;">
        <div style="font-size:2.8rem;font-family:'DM Serif Display',serif;color:#0f3460;line-height:1;">{aantal}</div>
        <div><div style="font-weight:600;color:#0f3460;font-size:1rem;">{"inzending ontvangen" if aantal==1 else "inzendingen ontvangen"}</div>
        <div style="color:#666;font-size:0.85rem;">Werkveld evaluaties</div></div></div>""", unsafe_allow_html=True)
    with c2:
        st.markdown(f"""<div style="background:#f0fdf4;border:1px solid #86efac;border-radius:12px;
        padding:1.2rem 1.8rem;display:flex;align-items:center;gap:1.2rem;">
        <div style="font-size:2.8rem;font-family:'DM Serif Display',serif;color:#166534;line-height:1;">{fg_aantal}</div>
        <div><div style="font-weight:600;color:#166534;font-size:1rem;">{"aanmelding" if fg_aantal==1 else "aanmeldingen"} focusgroep</div>
        <div style="color:#555;font-size:0.85rem;">Willen een Teams meeting</div></div></div>""", unsafe_allow_html=True)
    st.markdown("---")

    st.subheader("Niveauverdeling per onderdeel")
    an=[1,2,3,4,5]; kl=[NIVEAU_KLEUREN[n] for n in an]
    fig,axes=plt.subplots(1,len(oks),figsize=(7*len(oks),5),sharey=True); fig.patch.set_facecolor("#f8f9ff")
    if len(oks)==1: axes=[axes]
    for ax,key in zip(axes,oks):
        col=f"Niveau {key.split('-')[0].strip()}"
        if col in df.columns:
            counts=df[col].value_counts().sort_index(); cl=[counts.get(n,0) for n in an]
            bars=ax.bar(an,cl,color=kl,width=0.6,edgecolor="white",linewidth=1.5)
            for bar,cnt in zip(bars,cl):
                if cnt>0: ax.text(bar.get_x()+bar.get_width()/2,bar.get_height()+0.05,str(cnt),ha="center",va="bottom",fontsize=12,fontweight="bold",color="#1a1a2e")
            ax.set_title(key,fontsize=12,fontweight="bold",color="#0f3460",pad=10); ax.set_facecolor("#f8f9ff")
            ax.spines[["top","right"]].set_visible(False); ax.spines[["left","bottom"]].set_color("#ddd")
            ax.grid(axis="y",linestyle="--",alpha=0.35,color="#ccc"); ax.yaxis.set_major_locator(plt.MaxNLocator(integer=True)); ax.set_xticks(an)
    axes[0].set_ylabel("Aantal inzendingen",fontsize=10,color="#555")
    fig.suptitle("Niveauverdeling per Onderdeel",fontsize=14,fontweight="bold",color="#0f3460",y=1.03)
    plt.tight_layout(); st.pyplot(fig); plt.close(fig)
    st.markdown("---")

    st.subheader("Gemiddelde scores per vraag & per onderdeel")
    for sectie, vl in VRAGEN_WV.items():
        with st.expander(f"📌 {sectie}", expanded=False):
            gl=[r["scores"].get(sectie,0) for r in resultaten]
            gs=round(sum(gl)/len(gl),2) if gl else 0; ns=bereken_niveau(gs); ks=niveau_kleur_css(ns)
            st.markdown(f'<div style="background:{ks}22;border-left:4px solid {ks};border-radius:8px;padding:0.7rem 1rem;margin-bottom:1rem;"><strong>Onderdeel gemiddelde: {gs:.2f}</strong> &nbsp;->&nbsp; {NIVEAU_LABELS[ns]}</div>', unsafe_allow_html=True)
            rows=[]
            for vi,vraag in enumerate(vl,1):
                av=[r.get("scores_per_vraag",{}).get(sectie,[])[vi-1] for r in resultaten if vi-1<len(r.get("scores_per_vraag",{}).get(sectie,[]))]
                gv=round(sum(av)/len(av),2) if av else "-"
                rows.append({"Nr.":vi,"Vraag":vraag,"Gemiddelde":gv})
            st.dataframe(pd.DataFrame(rows),use_container_width=True,hide_index=True)
    st.markdown("---")

    st.subheader("Focusgroep aanmeldingen")
    fg_res=[r for r in resultaten if r.get("focusgroep")]
    if fg_res:
        st.dataframe(pd.DataFrame([{"E-mailadres":r.get("email",""),"Tijdstip":r.get("tijdstip","")} for r in fg_res]),use_container_width=True,hide_index=True)
    else:
        st.info("Nog geen aanmeldingen voor de focusgroep.")
    st.markdown("---")

    st.subheader("Resultaten downloaden")
    st.download_button("Download als Excel (.xlsx)", excel_werkveld(resultaten),
        f"werkveld_evaluatie_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
    st.markdown("---")
    if st.button("Verwijder alle werkveld resultaten"):
        verwijder_alle(TABEL_WV)
        st.success("Alle werkveld resultaten zijn verwijderd."); st.rerun()

# ═══════════════════════════════════════════════════════════════
#  DOCENT OMGEVING
# ═══════════════════════════════════════════════════════════════

# ═══════════════════════════════════════════════════════════════
#  DOCENT - eigen cursus evaluatie vragenlijst
# ═══════════════════════════════════════════════════════════════
def docent_evaluatie_pagina():
    st.markdown("""
    <div class="main-header">
        <h1>Docent Cursus Zelfevaluatie</h1>
        <p>Beoordeel de kwaliteit van uw eigen cursus aan de hand van onderstaande stellingen.<br>
        Elke stelling is gekoppeld aan een theoretisch kader uit de rubric.<br>
        <em>Kies per stelling: Niet aanwezig (1), Beperkt aanwezig (2), Naar wens aanwezig (3) of NVT.</em></p>
    </div>""", unsafe_allow_html=True)
    st.divider()

    alle_scores = {}
    alle_argumentaties = {}
    for onderdeel, stellingen in VRAGEN_DC.items():
        st.markdown(f'<div class="sectie-card"><h3>📌 {onderdeel}</h3></div>', unsafe_allow_html=True)
        sc = []
        args = []
        for i, (stelling, lens, theorie) in enumerate(stellingen, 1):
            col1, col2 = st.columns([3, 1])
            with col1:
                st.markdown(f"**{i}.** {stelling}")
                st.caption(f"Lens {lens} – {theorie}")
            with col2:
                score_optie = st.radio(
                    "Beoordeling",
                    options=["Niet aanwezig", "Beperkt aanwezig", "Naar wens aanwezig", "NVT"],
                    index=1,
                    key=f"dc_{onderdeel}_{i}",
                    label_visibility="collapsed",
                )
            argumentatie = st.text_area(
                "Toelichting (optioneel)",
                placeholder="Onderbouw uw beoordeling hier...",
                key=f"dc_arg_{onderdeel}_{i}",
                height=80,
                label_visibility="visible",
            )
            label_to_score = {
                "Niet aanwezig": 1,
                "Beperkt aanwezig": 2,
                "Naar wens aanwezig": 3,
                "NVT": None,
            }
            sc.append(label_to_score[score_optie])
            args.append(argumentatie.strip())
            st.markdown("---")
        alle_scores[onderdeel] = sc
        alle_argumentaties[onderdeel] = args

    c1, c2, c3 = st.columns([1, 2, 1])
    with c2:
        if st.button("Stuur mijn zelfevaluatie in", use_container_width=True):
            sg, sn = {}, {}
            for o, sc in alle_scores.items():
                # Exclude NVT (None) from average calculation
                geldig = [s for s in sc if s is not None]
                g = round(sum(geldig)/len(geldig), 2) if geldig else None
                sg[o] = g
                sn[o] = dc_label(g)
            geldig_totaal = [v for v in sg.values() if v is not None]
            tg = round(sum(geldig_totaal)/len(geldig_totaal), 2) if geldig_totaal else None
            sla_docent_op(alle_scores, sg, sn, tg, alle_argumentaties)
            st.session_state["dc_ingediend"] = True
            st.session_state["dc_resultaat"] = {"sectie_gemiddeldes": sg, "sectie_niveaus": sn, "totaal_gemiddelde": tg}
            st.rerun()


# ═══════════════════════════════════════════════════════════════
#  DOCENT - bedankt scherm na zelfevaluatie
# ═══════════════════════════════════════════════════════════════
def docent_evaluatie_bedankt():
    res = st.session_state.get("dc_resultaat", {})
    sg = res.get("sectie_gemiddeldes", {})
    sn = res.get("sectie_niveaus", {})
    tg = res.get("totaal_gemiddelde", 2)

    st.success("Uw zelfevaluatie is succesvol opgeslagen!")
    st.markdown("## Uw beoordeling per onderdeel")

    cols = st.columns(3)
    for idx, (o, gem) in enumerate(sg.items()):
        kleur = dc_kleur(gem)
        label = sn.get(o, "")
        with cols[idx % 3]:
            st.markdown(f'<div class="rubric-card" style="background:{kleur};color:#1a1a2e;"><h4>{o}</h4><div class="badge">{label}</div><div class="sub">Gemiddelde: {gem:.2f}</div></div>', unsafe_allow_html=True)

    kleur_tot = dc_kleur(tg)
    label_tot = dc_label(tg)
    gem_str = f"{tg:.2f}" if tg is not None else "–"
    st.markdown(f'<div class="totaal-badge" style="background:linear-gradient(135deg,{kleur_tot}cc,{kleur_tot});color:#1a1a2e;"><div class="label">Totaal Gemiddelde</div><div class="tekst">{label_tot}</div><div style="margin-top:0.5rem;opacity:0.85;font-size:0.9rem;">Gemiddelde over alle stellingen: {gem_str}</div></div>', unsafe_allow_html=True)

    st.markdown("&nbsp;")
    if st.button("Terug naar dashboard"):
        st.session_state["dc_ingediend"] = False
        st.session_state["dc_resultaat"] = {}
        st.session_state["docent_keuze"] = "Resultaten inzien"
        st.rerun()


# ═══════════════════════════════════════════════════════════════
#  DOCENT - dashboard zelfevaluatie resultaten
# ═══════════════════════════════════════════════════════════════
def dash_docent_evaluatie():
    st.markdown("""
    <div class="dashboard-header">
        <h2 style="margin:0;font-family:'DM Serif Display',serif;">Dashboard - Docent Zelfevaluatie</h2>
        <p style="margin:0.4rem 0 0;opacity:0.85;">Overzicht van alle ingediende docent zelfevaluaties. <br> 
        Graag na het downloaden van de resultaten de resultaten verwijderen zodat de volgende gerbuiker hier betrouwbare informatie uithaald.</p>
    </div>""", unsafe_allow_html=True)

    resultaten = laad(TABEL_DC)
    if not resultaten:
        st.warning("Er zijn nog geen docent zelfevaluaties ingestuurd.")
        return

    aantal = len(resultaten)
    st.markdown(f"""
    <div style="background:#eef2ff;border:1px solid #c7d2fe;border-radius:12px;
                padding:1.2rem 1.8rem;margin-bottom:1.5rem;display:flex;align-items:center;gap:1.2rem;">
        <div style="font-size:2.8rem;font-family:'DM Serif Display',serif;color:#0f3460;line-height:1;">{aantal}</div>
        <div><div style="font-weight:600;color:#0f3460;font-size:1rem;">{"evaluatie ingediend" if aantal==1 else "evaluaties ingediend"}</div>
        <div style="color:#666;font-size:0.85rem;">Docent zelfevaluaties</div></div>
    </div>""", unsafe_allow_html=True)

    # Gemiddelden per onderdeel
    st.subheader("Gemiddelde scores per onderdeel")
    for onderdeel in VRAGEN_DC.keys():
        gems = [r["sectie_gemiddeldes"].get(onderdeel) for r in resultaten if r["sectie_gemiddeldes"].get(onderdeel) is not None]
        gem = round(sum(gems)/len(gems), 2) if gems else None
        kleur = dc_kleur(gem)
        label = dc_label(gem)
        gem_str = f"{gem:.2f}" if gem is not None else "–"
        st.markdown(f'<div style="background:{kleur};border-radius:8px;padding:0.7rem 1rem;margin-bottom:0.5rem;"><strong>{onderdeel}:</strong> {label} (gemiddelde: {gem_str})</div>', unsafe_allow_html=True)

    st.markdown("---")

    # Gemiddelden per stelling
    st.subheader("Gemiddelde scores per stelling")
    for onderdeel, stellingen in VRAGEN_DC.items():
        with st.expander(f"📌 {onderdeel}", expanded=False):
            gl = [r["sectie_gemiddeldes"].get(onderdeel) for r in resultaten if r["sectie_gemiddeldes"].get(onderdeel) is not None]
            gs = round(sum(gl)/len(gl), 2) if gl else None
            kleur_s = dc_kleur(gs)
            gs_str = f"{gs:.2f}" if gs is not None else "–"
            st.markdown(f'<div style="background:{kleur_s};border-radius:8px;padding:0.7rem 1rem;margin-bottom:1rem;"><strong>Onderdeel gemiddelde: {gs_str}</strong></div>', unsafe_allow_html=True)
            rows = []
            for vi, (stelling, lens, theorie) in enumerate(stellingen, 1):
                raw_scores = [r.get("scores_per_stelling", {}).get(onderdeel, [])[vi-1]
                              for r in resultaten
                              if vi-1 < len(r.get("scores_per_stelling", {}).get(onderdeel, []))]
                # Filter out NVT (None values)
                geldig_sc = [s for s in raw_scores if s is not None]
                nvt_count = raw_scores.count(None)
                gv = round(sum(geldig_sc)/len(geldig_sc), 2) if geldig_sc else None
                gv_str = f"{gv:.2f}" if gv is not None else "–"
                label_v = dc_label(gv)
                nvt_str = f" ({nvt_count}× NVT)" if nvt_count > 0 else ""
                rows.append({"Nr.": vi, "Stelling": stelling, "Lens": lens, "Theorie": theorie, "Gemiddelde": gv_str + nvt_str, "Oordeel": label_v})
            st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

            # Toon argumentaties per stelling
            heeft_args = any(
                r.get("argumentaties", {}).get(onderdeel, [""] * len(stellingen))[vi-1]
                for r in resultaten
                for vi in range(1, len(stellingen)+1)
                if vi-1 < len(r.get("argumentaties", {}).get(onderdeel, []))
            )
            if heeft_args:
                st.markdown("**Toelichtingen per stelling:**")
                for vi, (stelling, lens, theorie) in enumerate(stellingen, 1):
                    args = [r.get("argumentaties", {}).get(onderdeel, [])[vi-1]
                            for r in resultaten
                            if vi-1 < len(r.get("argumentaties", {}).get(onderdeel, []))
                            and r.get("argumentaties", {}).get(onderdeel, [])[vi-1]]
                    if args:
                        st.markdown(f"*Stelling {vi}: {stelling[:80]}...*" if len(stelling)>80 else f"*Stelling {vi}: {stelling}*")
                        for idx_a, arg in enumerate(args, 1):
                            st.markdown(f'<div style="background:#f0f9ff;border-left:3px solid #7dd3fc;border-radius:6px;padding:0.5rem 0.8rem;margin-bottom:0.3rem;font-size:0.88rem;">#{idx_a}: {arg}</div>', unsafe_allow_html=True)

    st.markdown("---")

    # Rubric
    st.subheader("Rubric - Analysemodel Docent Zelfevaluatie")
    st.caption("Kleur geeft het gemiddelde aan over alle ingediende evaluaties. NVT-beoordelingen worden uitgesloten van het gemiddelde.")

    # Bereken per onderdeel, per theorie het gemiddelde (NVT = None uitgesloten)
    theorie_gems_dc = {}
    for o, stellingen in VRAGEN_DC.items():
        for i, (stelling, lens, theorie) in enumerate(stellingen):
            scores = []
            for r in resultaten:
                sc = r.get("scores_per_stelling", {}).get(o, [])
                if i < len(sc) and sc[i] is not None:
                    scores.append(sc[i])
            theorie_gems_dc[(o, theorie)] = round(sum(scores)/len(scores), 2) if scores else None

    lh_dc = [
        "Lens 1: Niveau & Samenhang<br><small>(Biggs, Bloom, Dublin)</small>",
        "Lens 2: Didactisch ontwerp<br><small>(Context-Concept, Entwistle, Gagné, Scaffolding)</small>",
        "Lens 3: Transfer theorie-praktijk<br><small>(Kolb, Miller, Shulman PCK)</small>",
        "Lens 4: De student en leertools<br><small>(TPACK, Zimmerman)</small>",
    ]

    st.markdown("""<div style="display:flex;gap:1rem;flex-wrap:wrap;margin-bottom:0.8rem;font-size:0.8rem;">
    <span style="background:#f8d7da;padding:3px 10px;border-radius:6px;">Rood = Niet aanwezig (&lt; 1.67)</span>
    <span style="background:#fff9c4;padding:3px 10px;border-radius:6px;">Geel = Beperkt aanwezig (1.67–2.33)</span>
    <span style="background:#d4edda;padding:3px 10px;border-radius:6px;">Groen = Naar wens aanwezig (&gt; 2.33)</span>
    <span style="background:#e8ecf4;padding:3px 10px;border-radius:6px;">Grijs = Geen data / NVT</span></div>""", unsafe_allow_html=True)

    tabel = '<table class="rubric-tabel" style="width:100%;border-collapse:collapse;"><thead><tr>'
    tabel += '<th style="background:#1e3a5f;color:white;padding:10px;width:16%;">Onderdeel</th>'
    for h in lh_dc:
        tabel += f'<th style="background:#1e3a5f;color:white;padding:10px;width:21%;">{h}</th>'
    tabel += "</tr></thead><tbody>"

    for o in VRAGEN_DC.keys():
        tabel += "<tr>"
        tabel += f'<td style="background:#eef2ff;font-weight:700;color:#0f3460;text-align:center;padding:10px;border:1px solid #ddd;vertical-align:top;">{o}</td>'

        for lens in [1,2,3,4]:
            theorieen_in_cel = []
            for stelling, sl, theorie in VRAGEN_DC[o]:
                if sl == lens:
                    theorieen_in_cel.append(theorie)
            gezien = set()
            unieke_theorieen = []
            for t in theorieen_in_cel:
                if t not in gezien:
                    gezien.add(t)
                    unieke_theorieen.append(t)

            if not unieke_theorieen:
                tabel += '<td style="background:#e8ecf4;border:1px solid #ccc;padding:10px;font-size:0.78rem;vertical-align:top;color:#888;">– (geen koppeling)</td>'
            else:
                cel_inhoud = ""
                for theorie in unieke_theorieen:
                    gem = theorie_gems_dc.get((o, theorie))
                    achter = dc_kleur(gem)
                    tekst_kleur = dc_kleur_tekst(gem) if gem is not None else "#555"
                    gem_str = f"{gem:.2f}" if gem is not None else "–"
                    label_str = dc_label(gem)

                    stelling_tekst = ""
                    for st_tekst, st_lens, st_theorie in VRAGEN_DC[o]:
                        if st_lens == lens and st_theorie == theorie:
                            stelling_tekst = st_tekst
                            break

                    cel_inhoud += f"""<div style="background:{achter};border-radius:6px;padding:7px 9px;margin-bottom:6px;">
                        <div style="font-weight:700;font-size:0.78rem;color:{tekst_kleur};margin-bottom:3px;">
                            {theorie} — {label_str} ({gem_str})
                        </div>
                        <div style="font-size:0.75rem;color:#333;line-height:1.4;">{stelling_tekst}</div>
                    </div>"""

                tabel += f'<td style="background:white;border:1px solid #ccc;padding:8px;vertical-align:top;">{cel_inhoud}</td>'

        tabel += "</tr>"
    tabel += "</tbody></table>"
    st.markdown(tabel, unsafe_allow_html=True)
    st.markdown("---")

    # Excel download
    st.subheader("Resultaten downloaden")
    excel_bytes = excel_docent(resultaten)
    st.download_button("Download als Excel (.xlsx)", excel_bytes,
        f"docent_evaluatie_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
    st.markdown("---")
    if st.button("Verwijder alle docent evaluatieresultaten"):
        verwijder_alle(TABEL_DC)
        st.success("Alle docent evaluatieresultaten zijn verwijderd.")
        st.rerun()


# ═══════════════════════════════════════════════════════════════
#  EXCEL EXPORT DOCENT
# ═══════════════════════════════════════════════════════════════
def excel_docent(resultaten):
    DB="0F3460"; LB="EEF2FF"; WIT="FFFFFF"
    thin=Side(style="thin",color="CCCCCC"); dik=Side(style="medium",color="0F3460")
    rand=Border(left=thin,right=thin,top=thin,bottom=thin)
    dik_rand=Border(left=dik,right=dik,top=dik,bottom=dik)
    wb=Workbook()

    # Blad 1: Samenvatting
    ws=wb.active; ws.title="Samenvatting"; ws.sheet_view.showGridLines=False
    ws.column_dimensions["A"].width=32
    for c in ["B","C","D"]: ws.column_dimensions[c].width=18
    ws.merge_cells("A1:D1"); ws["A1"]="Docent Zelfevaluatie - Resultaten Overzicht"
    ws["A1"].font=Font(name="Arial",bold=True,size=16,color=WIT); ws["A1"].fill=PatternFill("solid",fgColor=DB)
    ws["A1"].alignment=Alignment(horizontal="center",vertical="center"); ws.row_dimensions[1].height=36
    ws.merge_cells("A2:D2"); ws["A2"]=f"Gegenereerd: {datetime.now().strftime('%d-%m-%Y %H:%M')}  |  Totaal: {len(resultaten)}"
    ws["A2"].font=Font(name="Arial",size=10,color="666666"); ws["A2"].fill=PatternFill("solid",fgColor=LB)
    ws["A2"].alignment=Alignment(horizontal="center",vertical="center"); ws.row_dimensions[2].height=20; ws.row_dimensions[3].height=8
    for ci,h in enumerate(["Onderdeel","Gem. score","Oordeel","Responsen"],1):
        c=ws.cell(row=4,column=ci,value=h); c.font=Font(name="Arial",bold=True,size=10,color=WIT)
        c.fill=PatternFill("solid",fgColor=DB); c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); c.border=rand
    ws.row_dimensions[4].height=28
    KLEUR_DC = {"Naar wens aanwezig":"2ECC71","Beperkt aanwezig":"F1C40F","Niet aanwezig":"E74C3C","Geen data":"E8ECF4"}
    for rij,onderdeel in enumerate(VRAGEN_DC.keys(),5):
        gems=[r["sectie_gemiddeldes"].get(onderdeel) for r in resultaten if r["sectie_gemiddeldes"].get(onderdeel) is not None]
        gem=round(sum(gems)/len(gems),2) if gems else None
        label=dc_label(gem)
        gem_str=f"{gem:.2f}" if gem is not None else "–"
        kf=PatternFill("solid",fgColor=KLEUR_DC.get(label,"E8ECF4"))
        for ci,val in enumerate([onderdeel,gem_str,label,len(resultaten)],1):
            c=ws.cell(row=rij,column=ci,value=val); c.font=Font(name="Arial",size=10)
            c.alignment=Alignment(horizontal="center",vertical="center"); c.border=rand
            if ci==3: c.fill=kf; c.font=Font(name="Arial",size=10,bold=True,color=WIT if label=="Niet naar wens aanwezig" else "1a1a2e")
        ws.row_dimensions[rij].height=22

    # Blad 2: Gemiddelden per stelling
    ws2=wb.create_sheet("Gemiddelden per stelling"); ws2.sheet_view.showGridLines=False
    ws2.column_dimensions["A"].width=6; ws2.column_dimensions["B"].width=60; ws2.column_dimensions["C"].width=10; ws2.column_dimensions["D"].width=18; ws2.column_dimensions["E"].width=14; ws2.column_dimensions["F"].width=12
    ws2.merge_cells("A1:E1"); ws2["A1"]="Gemiddelde score per stelling"
    ws2["A1"].font=Font(name="Arial",bold=True,size=14,color=WIT); ws2["A1"].fill=PatternFill("solid",fgColor=DB)
    ws2["A1"].alignment=Alignment(horizontal="center",vertical="center"); ws2.row_dimensions[1].height=30
    rij=2
    for onderdeel,stellingen in VRAGEN_DC.items():
        ws2.row_dimensions[rij].height=8; rij+=1
        ws2.merge_cells(f"A{rij}:E{rij}"); ws2[f"A{rij}"]=onderdeel
        ws2[f"A{rij}"].font=Font(name="Arial",bold=True,size=11,color=WIT); ws2[f"A{rij}"].fill=PatternFill("solid",fgColor="1e3a5f")
        ws2[f"A{rij}"].alignment=Alignment(horizontal="left",vertical="center"); ws2[f"A{rij}"].border=rand; ws2.row_dimensions[rij].height=22; rij+=1
        for hi,h in enumerate(["#","Stelling","Lens","Theorie","Gem. score","NVT-count"],1):
            c=ws2.cell(row=rij,column=hi,value=h); c.font=Font(name="Arial",bold=True,size=9,color=WIT)
            c.fill=PatternFill("solid",fgColor=DB); c.alignment=Alignment(horizontal="center",vertical="center"); c.border=rand
        ws2.row_dimensions[rij].height=18; rij+=1
        for vi,(stelling,lens,theorie) in enumerate(stellingen,1):
            raw=[r.get("scores_per_stelling",{}).get(onderdeel,[])[vi-1] for r in resultaten if vi-1<len(r.get("scores_per_stelling",{}).get(onderdeel,[]))]
            geldig=[s for s in raw if s is not None]; nvt_cnt=raw.count(None)
            gv=round(sum(geldig)/len(geldig),2) if geldig else None
            gv_str=f"{gv:.2f}" if gv is not None else "–"
            kleur_cel=dc_kleur_hex(gv)
            kfv=PatternFill("solid",fgColor=kleur_cel)
            for ci,val in enumerate([vi,stelling,lens,theorie,gv_str,nvt_cnt if nvt_cnt>0 else ""],1):
                c=ws2.cell(row=rij,column=ci,value=val); c.font=Font(name="Arial",size=9)
                c.alignment=Alignment(horizontal="center" if ci!=2 else "left",vertical="center",wrap_text=(ci==2)); c.border=rand
                if ci==5: c.fill=kfv
            ws2.row_dimensions[rij].height=32; rij+=1

    # ── Blad 3: Rubric per theorie ──
    ws3 = wb.create_sheet("Rubric per theorie")
    ws3.sheet_view.showGridLines = False

    # Kolombreedtes: onderdeel + 4 lens-kolommen
    ws3.column_dimensions["A"].width = 24
    ws3.column_dimensions["B"].width = 42
    ws3.column_dimensions["C"].width = 42
    ws3.column_dimensions["D"].width = 42
    ws3.column_dimensions["E"].width = 42

    # Bereken per theorie het gemiddelde — NVT (None) uitgesloten
    theorie_gems = {}
    for o, stellingen in VRAGEN_DC.items():
        for i, (stelling, lens, theorie) in enumerate(stellingen):
            scores = []
            for r in resultaten:
                sc = r.get("scores_per_stelling", {}).get(o, [])
                if i < len(sc) and sc[i] is not None:
                    scores.append(sc[i])
            theorie_gems[(o, theorie)] = round(sum(scores)/len(scores), 2) if scores else None

    KLEUR_GOED    = "D4EDDA"
    KLEUR_NEUTR   = "FFF9C4"
    KLEUR_SLECHT  = "F8D7DA"
    KLEUR_GEEN    = "E8ECF4"

    def cel_kleur(gem):
        if gem is None: return KLEUR_GEEN
        if gem >= 2.34: return KLEUR_GOED
        elif gem >= 1.67: return KLEUR_NEUTR
        else: return KLEUR_SLECHT

    def cel_label(gem):
        if gem is None: return "Geen data / NVT"
        if gem >= 2.34: return "Naar wens aanwezig"
        elif gem >= 1.67: return "Beperkt aanwezig"
        else: return "Niet aanwezig"

    # Titelrij
    ws3.merge_cells("A1:E1")
    ws3["A1"] = "Rubric Docent Zelfevaluatie – Score per theorie"
    ws3["A1"].font = Font(name="Arial", bold=True, size=14, color=WIT)
    ws3["A1"].fill = PatternFill("solid", fgColor=DB)
    ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 34

    ws3.merge_cells("A2:E2")
    ws3["A2"] = "Groen = Naar wens aanwezig (>= 2.34)  |  Geel = Beperkt aanwezig (1.67–2.33)  |  Rood = Niet aanwezig (< 1.67)  |  Grijs = Geen data / NVT"
    ws3["A2"].font = Font(name="Arial", size=9, italic=True, color="444444")
    ws3["A2"].fill = PatternFill("solid", fgColor=LB)
    ws3["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[2].height = 18
    ws3.row_dimensions[3].height = 6

    # Kolomhoofden
    lens_headers = [
        "Lens 1: Niveau & Samenhang\n(Biggs, Bloom, Dublin)",
        "Lens 2: Didactisch ontwerp\n(Context-Concept, Entwistle, Gagne, Scaffolding)",
        "Lens 3: Transfer theorie-praktijk\n(Kolb, Miller, Shulman PCK)",
        "Lens 4: De student en leertools\n(TPACK, Zimmerman)",
    ]
    for ci, h in enumerate(["Onderdeel"] + lens_headers, 1):
        c = ws3.cell(row=4, column=ci, value=h)
        c.font = Font(name="Arial", bold=True, size=9, color=WIT)
        c.fill = PatternFill("solid", fgColor="1e3a5f")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = rand
    ws3.row_dimensions[4].height = 44

    # Data rijen — per onderdeel één rij, per lens de theorieën met elk hun kleur
    for ri, (onderdeel, stellingen) in enumerate(VRAGEN_DC.items(), start=5):
        # Bepaal rijhoogte op basis van max theorieën per lens
        max_theorieen = max(
            len([s for s in stellingen if s[1] == l]) for l in [1,2,3,4]
        )
        ws3.row_dimensions[ri].height = max(60, max_theorieen * 45)

        # Onderdeel naam
        c = ws3.cell(row=ri, column=1, value=onderdeel)
        c.font = Font(name="Arial", bold=True, size=10)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = dik_rand
        c.fill = PatternFill("solid", fgColor=LB)

        # Per lens
        for lens in [1, 2, 3, 4]:
            # Verzamel alle stellingen voor dit onderdeel + lens
            cel_stellingen = [(st, th) for st, sl, th in stellingen if sl == lens]

            if not cel_stellingen:
                c = ws3.cell(row=ri, column=lens + 1, value="– (geen koppeling)")
                c.font = Font(name="Arial", size=8, color="888888", italic=True)
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border = rand
                c.fill = PatternFill("solid", fgColor=KLEUR_GEEN)
            else:
                # Bouw celinhoud: naam theorie + label + gem + stellingtekst
                cel_tekst = ""
                for stelling_tekst, theorie in cel_stellingen:
                    gem = theorie_gems.get((onderdeel, theorie))
                    label = cel_label(gem)
                    gem_str = f"{gem:.2f}" if gem is not None else "–"
                    cel_tekst += f"{theorie} - {label} ({gem_str})\n{stelling_tekst}\n\n"

                c = ws3.cell(row=ri, column=lens + 1, value=cel_tekst.strip())
                c.font = Font(name="Arial", size=8)
                c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                c.border = rand

                # Celkleur: gemiddelde van alle theorieën in deze cel
                gems_in_cel = [theorie_gems.get((onderdeel, th)) for _, th in cel_stellingen
                               if theorie_gems.get((onderdeel, th)) is not None]
                gem_cel = round(sum(gems_in_cel)/len(gems_in_cel), 2) if gems_in_cel else None
                c.fill = PatternFill("solid", fgColor=cel_kleur(gem_cel))

    # Blad 4: Argumentaties per stelling
    ws4 = wb.create_sheet("Toelichtingen")
    ws4.sheet_view.showGridLines = False
    ws4.column_dimensions["A"].width = 6
    ws4.column_dimensions["B"].width = 28
    ws4.column_dimensions["C"].width = 55
    ws4.column_dimensions["D"].width = 10
    ws4.column_dimensions["E"].width = 50
    ws4.merge_cells("A1:E1"); ws4["A1"] = "Toelichtingen per stelling (open argumentaties)"
    ws4["A1"].font = Font(name="Arial", bold=True, size=14, color=WIT)
    ws4["A1"].fill = PatternFill("solid", fgColor=DB)
    ws4["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws4.row_dimensions[1].height = 30
    for ci, h in enumerate(["#","Onderdeel","Stelling","Lens","Toelichting"],1):
        c = ws4.cell(row=2, column=ci, value=h)
        c.font = Font(name="Arial", bold=True, size=9, color=WIT)
        c.fill = PatternFill("solid", fgColor=DB)
        c.alignment = Alignment(horizontal="center", vertical="center"); c.border = rand
    ws4.row_dimensions[2].height = 18
    arg_rij = 3; arg_nr = 1
    for o, stellingen in VRAGEN_DC.items():
        for vi, (stelling, lens, theorie) in enumerate(stellingen):
            args = [r.get("argumentaties", {}).get(o, [])[vi]
                    for r in resultaten
                    if vi < len(r.get("argumentaties", {}).get(o, []))
                    and r.get("argumentaties", {}).get(o, [])[vi]]
            for arg_tekst in args:
                achter = "F8F9FF" if arg_rij % 2 == 0 else "FFFFFF"
                for ci, val in enumerate([arg_nr, o, stelling, lens, arg_tekst], 1):
                    c = ws4.cell(row=arg_rij, column=ci, value=val)
                    c.font = Font(name="Arial", size=9)
                    c.alignment = Alignment(horizontal="center" if ci not in (2,3,5) else "left",
                                            vertical="top", wrap_text=True)
                    c.border = rand; c.fill = PatternFill("solid", fgColor=achter)
                ws4.row_dimensions[arg_rij].height = 40
                arg_rij += 1; arg_nr += 1
    if arg_rij == 3:
        ws4.merge_cells("A3:E3"); ws4["A3"] = "Geen toelichtingen ingestuurd."
        ws4["A3"].font = Font(name="Arial", size=10, italic=True, color="888888")
        ws4["A3"].alignment = Alignment(horizontal="center", vertical="center")

    buf=io.BytesIO(); wb.save(buf); buf.seek(0); return buf.getvalue()

def docent_login():
    if st.button("<- Terug naar startpagina"):
        st.session_state["rol"] = None; st.rerun()
    st.markdown('<div class="main-header"><h1>Docentendashboard</h1><p>Voer het wachtwoord in om toegang te krijgen.</p></div>', unsafe_allow_html=True)
    ww = st.text_input("Wachtwoord", type="password", placeholder="••••••••")
    if st.button("Inloggen"):
        if ww == DOCENT_WACHTWOORD: st.session_state["docent_ingelogd"] = True; st.rerun()
        else: st.error("Onjuist wachtwoord.")

def docent_omgeving():
    c1, c2, c3, c4 = st.columns([2, 2, 2, 1])
    with c1:
        st.markdown('<div style="color:#0f3460;font-weight:700;font-size:1.05rem;padding-top:0.5rem;">Docentendashboard</div>', unsafe_allow_html=True)
    with c2:
        hoofd_keuze = st.selectbox("Kies optie:",
                             ["Resultaten inzien", "Eigen cursus evalueren"],
                             key="docent_hoofd_keuze", label_visibility="collapsed")
    with c3:
        if hoofd_keuze == "Resultaten inzien":
            dash_keuze = st.selectbox("Bekijk resultaten van:",
                                 ["Studenten", "Werkveld / Stagebegeleiders", "Docent Zelfevaluatie"],
                                 key="dash_keuze", label_visibility="collapsed")
        else:
            dash_keuze = None
    with c4:
        if st.button("Uitloggen", use_container_width=True):
            st.session_state["docent_ingelogd"] = False; st.rerun()
    st.divider()

    if hoofd_keuze == "Eigen cursus evalueren":
        if st.session_state.get("dc_ingediend", False):
            docent_evaluatie_bedankt()
        else:
            docent_evaluatie_pagina()
    else:
        if dash_keuze == "Studenten":
            dash_studenten()
        elif dash_keuze == "Werkveld / Stagebegeleiders":
            dash_werkveld()
        else:
            dash_docent_evaluatie()

# ═══════════════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════════════
def main():
    st.set_page_config(page_title="Cursusevaluatie ALO", page_icon="📋", layout="wide")
    laad_stijl()
    defaults = [
        ("rol", None), ("docent_ingelogd", False),
        ("wv_email", None), ("wv_sh_gezien", False),
        ("wv_ingediend", False), ("wv_resultaat", {}),
        ("st_ingediend", False), ("st_resultaat", {}),
        ("dc_ingediend", False), ("dc_resultaat", {}),
        ("docent_keuze", "Resultaten inzien"),
    ]
    for k, v in defaults:
        if k not in st.session_state:
            st.session_state[k] = v
    rol = st.session_state["rol"]
    if rol is None:
        landingspagina()
    elif rol == "student":
        if st.session_state["st_ingediend"]:
            st_bedankt()
        else:
            student_pagina()
    elif rol == "werkveld":
        if st.session_state["wv_ingediend"]:
            wv_bedankt()
        elif st.session_state["wv_email"] is None:
            wv_email_scherm()
        elif not st.session_state["wv_sh_gezien"]:
            wv_studiehandleiding_scherm()
        else:
            wv_vragenlijst()
    elif rol == "docent":
        if st.session_state["docent_ingelogd"]:
            docent_omgeving()
        else:
            docent_login()

if __name__ == "__main__":
    main()
